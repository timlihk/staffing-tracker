#!/usr/bin/env python3
"""
Parse Excel file to detect strikethrough formatting in Fee Arrangement column
and mark corresponding milestones as completed in the database.
"""

import openpyxl
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
import psycopg2
from datetime import datetime
import os
import re

# Database connection
DATABASE_URL = os.environ.get('DATABASE_URL', 'postgresql://postgres:qtSnaaSaelqHVydazTmViwejXbkkZxVY@crossover.proxy.rlwy.net:15782/railway')

# Excel file path
EXCEL_FILE = os.environ.get(
    'EXCEL_FILE',
    '/home/timlihk/staffing-tracker/billing-matter/HKCM Project List(81764217.1)_6Oct25.xlsx'
)

def connect_db():
    """Connect to PostgreSQL database"""
    return psycopg2.connect(DATABASE_URL)

def has_strikethrough(cell):
    """
    Check if cell has strikethrough formatting.
    Returns list of strikethrough text segments.
    """
    strikethrough_texts = []

    # Check if cell value is a CellRichText object (contains multiple formatting)
    if isinstance(cell.value, CellRichText):
        for text_block in cell.value:
            if isinstance(text_block, TextBlock):
                # Check if this text block has strikethrough
                if text_block.font and text_block.font.strike:
                    strikethrough_texts.append(text_block.text)
    # Check if cell has simple font formatting with strikethrough
    elif cell.font and cell.font.strike and cell.value:
        strikethrough_texts.append(str(cell.value))

    return strikethrough_texts

def extract_milestone_identifier(text):
    """
    Extract milestone identifier from text (e.g., '(a)', '(b)', '(c)')
    """
    match = re.search(r'\(([a-z])\)', text, re.IGNORECASE)
    return match.group(1).lower() if match else None

def parse_excel_for_strikethroughs():
    """
    Parse Excel file to find strikethrough text in Fee Arrangement column
    Returns dict mapping project_name -> list of strikethrough milestone identifiers
    """
    print(f"Loading Excel file: {EXCEL_FILE}", flush=True)
    print("This may take 2-3 minutes due to file size...", flush=True)
    import time
    start_time = time.time()

    workbook = openpyxl.load_workbook(EXCEL_FILE, data_only=False, keep_vba=False)

    load_time = time.time() - start_time
    print(f"✅ Loaded in {load_time:.1f} seconds", flush=True)

    # Use the Transactions worksheet (the first/active one)
    sheet = workbook['Transactions']

    print(f"Worksheet: {sheet.title}", flush=True)

    # Headers are in row 4 (not row 1)
    HEADER_ROW = 4
    headers = {}
    for cell in sheet[HEADER_ROW]:
        if cell.value:
            header_name = str(cell.value).strip()
            headers[header_name] = cell.column
            print(f"  Col {openpyxl.utils.get_column_letter(cell.column)}: {header_name}", flush=True)

    print(f"\nFound {len(headers)} headers", flush=True)

    # Find relevant columns - use exact header names from row 4
    project_name_col = headers.get('Project Name')
    cm_no_col = headers.get('C/M No')
    fee_arrangement_col = headers.get('Fee arrangement')
    if not fee_arrangement_col:
        # Newer HKCM files keep fee text in column I with "(a)" sub-header.
        fee_arrangement_col = headers.get('(a)')

    if not project_name_col or not fee_arrangement_col:
        print("ERROR: Could not find required columns", flush=True)
        print(f"Project Name column: {project_name_col}", flush=True)
        print(f"Fee Arrangement column: {fee_arrangement_col}", flush=True)
        return {}

    print(f"Project Name column: {openpyxl.utils.get_column_letter(project_name_col)}", flush=True)
    print(f"C/M No column: {openpyxl.utils.get_column_letter(cm_no_col)}", flush=True)
    print(f"Fee Arrangement column: {openpyxl.utils.get_column_letter(fee_arrangement_col)}", flush=True)

    # Data starts from row 5 (row 4 is headers)
    DATA_START_ROW = 5
    print(f"\nScanning rows {DATA_START_ROW} to {sheet.max_row} for strikethrough formatting...", flush=True)

    # Parse each row
    completed_milestones = {}

    for row_idx in range(DATA_START_ROW, sheet.max_row + 1):
        project_name_cell = sheet.cell(row_idx, project_name_col)
        project_name = str(project_name_cell.value).strip() if project_name_cell.value else None

        if not project_name:
            continue

        # Get CM number if available
        cm_no = None
        if cm_no_col:
            cm_no_cell = sheet.cell(row_idx, cm_no_col)
            cm_no = str(cm_no_cell.value).strip() if cm_no_cell.value else None

        # Check fee arrangement cell for strikethrough
        fee_cell = sheet.cell(row_idx, fee_arrangement_col)
        strikethrough_texts = has_strikethrough(fee_cell)

        if strikethrough_texts:
            print(f"\n=== Row {row_idx}: {project_name} (C/M: {cm_no}) ===")
            print(f"Found strikethrough text:")

            milestone_ids = []
            for text in strikethrough_texts:
                print(f"  - {text[:100]}...")

                # Extract milestone identifier (a, b, c, etc.)
                milestone_id = extract_milestone_identifier(text)
                if milestone_id:
                    milestone_ids.append(milestone_id)
                    print(f"    -> Milestone: ({milestone_id})")

            if milestone_ids:
                completed_milestones[project_name] = {
                    'milestones': milestone_ids,
                    'cm_no': cm_no
                }

    workbook.close()
    return completed_milestones

def update_database(completed_milestones):
    """
    Update database to mark milestones as completed
    """
    if not completed_milestones:
        print("\nNo completed milestones found with strikethrough formatting")
        return

    conn = connect_db()
    cursor = conn.cursor()

    completion_date = datetime.now()
    completion_source = 'excel_strikethrough'

    print(f"\n=== Updating Database ===")
    print(f"Found {len(completed_milestones)} projects with completed milestones")

    total_updated = 0

    for project_name, data in completed_milestones.items():
        milestone_ids = data['milestones']
        cm_no = data['cm_no']

        print(f"\nProject: {project_name}")
        print(f"  C/M No: {cm_no}")
        print(f"  Completed milestones: {milestone_ids}")

        # Find project in database
        cursor.execute("""
            SELECT bp.project_id, bp.project_name
            FROM billing_project bp
            WHERE bp.project_name = %s
            LIMIT 1
        """, (project_name,))

        project = cursor.fetchone()

        if not project:
            print(f"  ⚠️  Project not found in database")
            continue

        project_id = project[0]

        # Update each milestone
        for milestone_ordinal in milestone_ids:
            # milestone_ordinal is just 'a', 'b', 'c' - need to match against ordinal column which is '(a)', '(b)', etc.
            ordinal_value = f'({milestone_ordinal})'

            cursor.execute("""
                UPDATE billing_milestone
                SET
                    completed = TRUE,
                    completion_date = %s,
                    completion_source = %s
                WHERE fee_id IN (
                    SELECT bfa.fee_id
                    FROM billing_fee_arrangement bfa
                    JOIN billing_engagement be ON be.engagement_id = bfa.engagement_id
                    WHERE be.project_id = %s
                )
                AND ordinal = %s
                AND completed = FALSE
            """, (completion_date, completion_source, project_id, ordinal_value))

            updated_count = cursor.rowcount

            if updated_count > 0:
                print(f"  ✅ Marked milestone {ordinal_value} as completed")
                total_updated += updated_count
            else:
                print(f"  ⚠️  Milestone {ordinal_value} not found or already completed")

    conn.commit()
    cursor.close()
    conn.close()

    print(f"\n=== Summary ===")
    print(f"Total milestones marked as completed: {total_updated}")

def main():
    """Main function"""
    print("=" * 60)
    print("Parsing Excel for Strikethrough Milestones")
    print("=" * 60)

    # Parse Excel file
    completed_milestones = parse_excel_for_strikethroughs()

    # Update database
    update_database(completed_milestones)

    print("\n✅ Script completed successfully")

if __name__ == '__main__':
    main()
