#!/usr/bin/env python3
"""
Update Billing Database with Kimi-Powered Milestone Parser

Updates:
- billing_project (project info)
- billing_project_cm_no (CM numbers and financials)
- billing_engagement (engagements)
- billing_fee_arrangement (fee text)
- billing_milestone (parsed milestones)

Usage:
    python3 update-billing-kimi.py --dry-run    # Preview
    python3 update-billing-kimi.py              # Apply
    python3 update-billing-kimi.py --row 5      # Specific row
"""

import os
import sys
import re
import argparse
import zipfile
import xml.etree.ElementTree as ET
from typing import Dict, List, Optional, Set, Any, Tuple
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal

import openpyxl
import psycopg2
from psycopg2.extras import RealDictCursor

# Configuration
DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://postgres:qtSnaaSaelqHVydazTmViwejXbkkZxVY@crossover.proxy.rlwy.net:15782/railway",
)

DEFAULT_EXCEL_FILE = "/Users/timli/Library/CloudStorage/OneDrive-Personal/Coding/staffing-tracker/Billing/HKCM Project List (2026.02.12).xlsx"
EXCEL_FILE = os.environ.get("EXCEL_FILE", DEFAULT_EXCEL_FILE)

NS = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}


@dataclass
class Milestone:
    """Represents a parsed milestone"""
    ordinal: str
    description: str
    amount: Optional[float] = None
    currency: str = "USD"
    percent: Optional[float] = None
    is_conditional: bool = False
    completed: bool = False
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            'ordinal': self.ordinal,
            'description': self.description,
            'amount': self.amount,
            'currency': self.currency,
            'percent': self.percent,
            'is_conditional': self.is_conditional,
            'completed': self.completed,
        }


@dataclass
class ProjectData:
    """Complete project data from Excel"""
    row_num: int
    project_name: str
    client_name: str
    cm_no: str
    attorney_in_charge: str
    sca: str
    fees_usd: float
    billing_usd: float
    collection_usd: float
    billing_credit_usd: float
    ubt_usd: float
    ar_usd: float
    billing_credit_cny: float
    ubt_cny: float
    finance_comment: str
    remarks: str
    matter_notes: str
    fee_arrangement_text: str = ""
    milestones: List[Milestone] = field(default_factory=list)
    completed_milestones: Set[str] = field(default_factory=set)


class KimiMilestoneParser:
    """Kimi-style intelligent milestone parser"""
    
    CURRENCY_PATTERNS = [
        (r'US\$|\$|USD|美元|美金', 'USD'),
        (r'RMB|CNY|￥|人民币|元', 'CNY'),
        (r'HK\$|HKD|港元|港币', 'HKD'),
    ]
    
    ORDINAL_PATTERNS = [
        (r'\(([a-z])\)', 'letter'),
        (r'\((\d{1,2}[a-z]?)\)', 'number'),
        (r'^(\d+)\.', 'number_cn'),
    ]
    
    CONDITIONAL_KEYWORDS = [
        '减去', '扣除', '减去本所', '按', '根据', '实际产生',
        '若', '如果', '假如', '未能', '成功',
        'minus', 'deduct', 'based on', 'actual', 'if', 'condition'
    ]
    
    def detect_currency(self, text: str) -> str:
        text_upper = text.upper()
        for pattern, currency in self.CURRENCY_PATTERNS:
            if re.search(pattern, text_upper):
                return currency
        return "USD"
    
    def is_conditional(self, text: str) -> bool:
        text_lower = text.lower()
        return any(kw in text_lower for kw in self.CONDITIONAL_KEYWORDS)
    
    def extract_amount(self, text: str) -> Tuple[Optional[float], bool]:
        is_conditional = self.is_conditional(text)
        text_clean = text.replace(',', '').replace('，', '')
        
        patterns = [
            r'[US\$\$]\s*([\d]+(?:\.\d+)?)',
            r'([\d]+(?:\.\d+)?)\s*(?:USD|美元|美金)',
            r'(?:CNY|RMB|￥)\s*([\d]+(?:\.\d+)?)',
            r'([\d]+(?:\.\d+)?)\s*(?:元|人民币)',
            r'[-–—]\s*([\d]+(?:\.\d+)?)\s*$',
            r'\b([\d]{4,}(?:\.\d+)?)\b',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text_clean)
            if match:
                amount_str = match.group(1)
                if amount_str:
                    try:
                        amount = float(amount_str)
                        if amount > 1000:
                            return amount, is_conditional
                    except ValueError:
                        continue
        
        return None, is_conditional
    
    def extract_percentage(self, text: str) -> Optional[float]:
        patterns = [
            r'\((\d+(?:\.\d+)?)%\)',
            r'(\d+(?:\.\d+)?)%',
            r'（(\d+(?:\.\d+)?)%）',
        ]
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                try:
                    return float(match.group(1))
                except ValueError:
                    continue
        return None
    
    def find_ordinals(self, text: str) -> List[Tuple[str, int, int, str]]:
        ordinals = []
        seen = set()
        
        for pattern, otype in self.ORDINAL_PATTERNS:
            for match in re.finditer(pattern, text, re.IGNORECASE):
                ordinal = f"({match.group(1).lower()})"
                if ordinal not in seen:
                    seen.add(ordinal)
                    ordinals.append((ordinal, match.start(), match.end(), otype))
        
        ordinals.sort(key=lambda x: x[1])
        return ordinals
    
    def parse(self, text: str, struck_ordinals: Set[str] = None) -> List[Milestone]:
        if not text or not text.strip():
            return []
        
        if struck_ordinals is None:
            struck_ordinals = set()
        
        ordinals = self.find_ordinals(text)
        
        if not ordinals:
            amount, is_conditional = self.extract_amount(text)
            if amount:
                return [Milestone(
                    ordinal="(1)",
                    description=text.strip()[:200],
                    amount=amount,
                    currency=self.detect_currency(text),
                    is_conditional=is_conditional,
                    completed="*" in struck_ordinals
                )]
            return []
        
        milestones = []
        for i, (ordinal, start, end, otype) in enumerate(ordinals):
            if i + 1 < len(ordinals):
                segment_end = ordinals[i + 1][1]
            else:
                segment_end = len(text)
            
            segment = text[end:segment_end].strip()
            segment = re.sub(r'\n\s*\([a-z0-9]\).*$', '', segment, flags=re.DOTALL)
            
            amount, is_conditional = self.extract_amount(segment)
            currency = self.detect_currency(segment)
            percent = self.extract_percentage(segment)
            completed = ordinal in struck_ordinals or "*" in struck_ordinals
            
            milestones.append(Milestone(
                ordinal=ordinal,
                description=segment[:200],
                amount=amount,
                currency=currency,
                percent=percent,
                is_conditional=is_conditional,
                completed=completed
            ))
        
        return milestones


class ExcelStrikethroughExtractor:
    """Extract strikethrough from Excel XML"""
    
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.ss_list = []
        self._load_shared_strings()
    
    def _load_shared_strings(self):
        with zipfile.ZipFile(self.excel_path, 'r') as zf:
            if 'xl/sharedStrings.xml' not in zf.namelist():
                return
            
            ss_xml = zf.read('xl/sharedStrings.xml')
            root = ET.fromstring(ss_xml)
            
            for idx, si in enumerate(root.findall('.//main:si', NS)):
                parts = []
                has_strike = False
                for r in si.findall('.//main:r', NS):
                    t = r.find('.//main:t', NS)
                    text = t.text if t is not None else ""
                    rpr = r.find('.//main:rPr', NS)
                    struck = False
                    if rpr is not None:
                        strike = rpr.find('.//main:strike', NS)
                        if strike is not None:
                            val = strike.get('val', 'true')
                            if val in ('true', '1', 'single'):
                                struck = True
                                has_strike = True
                    parts.append({'text': text, 'struck': struck})
                
                full_text = ''.join(p['text'] for p in parts)
                self.ss_list.append({
                    'index': idx,
                    'parts': parts,
                    'has_strike': has_strike,
                    'full_text': full_text
                })
    
    def extract_for_cell(self, fee_text: str, cell_font_strike: bool = False) -> Set[str]:
        if cell_font_strike:
            return {"*"}
        
        struck_ordinals = set()
        
        for ss in self.ss_list:
            if not ss['has_strike']:
                continue
            
            # Check if this shared string matches our text
            if fee_text and len(fee_text) > 20:
                fee_start = fee_text[:50].replace('\n', ' ')
                ss_start = ss['full_text'][:50].replace('\n', ' ')
                
                if fee_start in ss_start or ss_start in fee_start or self._similar(fee_start, ss_start):
                    for part in ss['parts']:
                        if part['struck']:
                            for match in re.finditer(r'\(([a-z])\)', part['text'], re.I):
                                struck_ordinals.add(f"({match[1].lower()})")
        
        return struck_ordinals
    
    def _similar(self, a: str, b: str) -> bool:
        """Quick similarity check"""
        if len(a) < 20 or len(b) < 20:
            return False
        # Check if first 30 chars are similar
        return a[:30] in b or b[:30] in a


class DatabaseUpdater:
    """Handle all database updates"""
    
    def __init__(self, db_url: str, dry_run: bool = False):
        self.db_url = db_url
        self.dry_run = dry_run
        self.conn = None
        self.stats = {
            'projects_updated': 0,
            'cm_no_updated': 0,
            'engagements_created': 0,
            'fee_arrangements_created': 0,
            'milestones_inserted': 0,
            'milestones_updated': 0,
            'errors': []
        }
    
    def connect(self):
        if not self.dry_run:
            self.conn = psycopg2.connect(self.db_url)
            self.conn.autocommit = False
    
    def close(self):
        if self.conn:
            self.conn.close()
    
    def get_project_by_cm(self, cm_no: str) -> Optional[int]:
        """Find project by CM number, or None if not found"""
        if not cm_no or cm_no.strip() == '' or cm_no.strip().lower() == 'none':
            return None
        
        with self.conn.cursor() as cur:
            # Look up CM number in billing_project_cm_no
            cur.execute(
                """SELECT p.project_id 
                   FROM billing_project p
                   JOIN billing_project_cm_no c ON p.project_id = c.project_id
                   WHERE c.cm_no = %s""",
                (cm_no,)
            )
            result = cur.fetchone()
            if result:
                return result[0]
            
            # Try partial match on CM number
            cur.execute(
                """SELECT p.project_id 
                   FROM billing_project p
                   JOIN billing_project_cm_no c ON p.project_id = c.project_id
                   WHERE c.cm_no ILIKE %s""",
                (f"%{cm_no}%",)
            )
            result = cur.fetchone()
            if result:
                return result[0]
        
        return None
    
    def get_project_by_name(self, project_name: str, client_name: str) -> Optional[int]:
        """Find project by name as fallback"""
        if not project_name or project_name.strip() == '' or project_name.strip().lower() == 'none':
            return None
        
        with self.conn.cursor() as cur:
            # Try exact match first
            cur.execute(
                "SELECT project_id FROM billing_project WHERE project_name = %s",
                (project_name,)
            )
            result = cur.fetchone()
            if result:
                return result[0]
            
            # Try partial match
            cur.execute(
                "SELECT project_id FROM billing_project WHERE project_name ILIKE %s",
                (f"%{project_name}%",)
            )
            result = cur.fetchone()
            if result:
                return result[0]
            
            # Try client_name match
            if client_name and client_name.strip() != '' and client_name.strip().lower() != 'none':
                cur.execute(
                    "SELECT project_id FROM billing_project WHERE client_name ILIKE %s",
                    (f"%{client_name}%",)
                )
                result = cur.fetchone()
                if result:
                    return result[0]
        
        return None
    
    def update_project(self, project_id: int, data: ProjectData):
        """Update project basic info"""
        with self.conn.cursor() as cur:
            cur.execute("""
                UPDATE billing_project SET
                    client_name = %s,
                    attorney_in_charge = %s,
                    sca = %s,
                    updated_at = NOW()
                WHERE project_id = %s
            """, (
                data.client_name, data.attorney_in_charge, data.sca,
                project_id
            ))
            self.stats['projects_updated'] += 1
    
    def get_or_create_cm_no(self, project_id: int, cm_no: str, data: ProjectData) -> int:
        """Get or create CM number entry, returns cm_id"""
        with self.conn.cursor() as cur:
            # Check if exists
            cur.execute(
                "SELECT cm_id FROM billing_project_cm_no WHERE cm_no = %s",
                (cm_no,)
            )
            result = cur.fetchone()
            
            if result:
                cm_id = result[0]
                # Update financials
                cur.execute("""
                    UPDATE billing_project_cm_no SET
                        project_id = %s,
                        billing_to_date_usd = %s,
                        collected_to_date_usd = %s,
                        ubt_usd = %s,
                        billing_credit_usd = %s,
                        billing_credit_cny = %s,
                        ubt_cny = %s,
                        financials_updated_at = NOW()
                    WHERE cm_id = %s
                """, (
                    project_id,
                    data.billing_usd,
                    data.collection_usd,
                    data.ubt_usd,
                    data.billing_credit_usd,
                    data.billing_credit_cny,
                    data.ubt_cny,
                    cm_id
                ))
            else:
                # Create new
                cur.execute("""
                    INSERT INTO billing_project_cm_no (
                        project_id, cm_no, is_primary, status,
                        billing_to_date_usd, collected_to_date_usd,
                        ubt_usd, billing_credit_usd,
                        billing_credit_cny, ubt_cny,
                        financials_updated_at
                    ) VALUES (%s, %s, FALSE, 'active', %s, %s, %s, %s, %s, %s, NOW())
                    RETURNING cm_id
                """, (
                    project_id, cm_no,
                    data.billing_usd, data.collection_usd,
                    data.ubt_usd, data.billing_credit_usd,
                    data.billing_credit_cny, data.ubt_cny
                ))
                cm_id = cur.fetchone()[0]
            
            self.stats['cm_no_updated'] += 1
            return cm_id
    
    def get_or_create_engagement(self, project_id: int, cm_id: int, data: ProjectData) -> int:
        """Get or create engagement, returns engagement_id"""
        with self.conn.cursor() as cur:
            # Look for existing engagement
            cur.execute(
                "SELECT engagement_id FROM billing_engagement WHERE cm_id = %s",
                (cm_id,)
            )
            result = cur.fetchone()
            
            if result:
                return result[0]
            
            # Create new engagement
            cur.execute("""
                INSERT INTO billing_engagement (
                    project_id, cm_id, engagement_code, engagement_title, updated_at
                ) VALUES (%s, %s, 'original', %s, NOW())
                RETURNING engagement_id
            """, (project_id, cm_id, data.project_name[:100]))
            
            engagement_id = cur.fetchone()[0]
            self.stats['engagements_created'] += 1
            return engagement_id
    
    def get_or_create_fee_arrangement(self, engagement_id: int, data: ProjectData) -> int:
        """Get or create fee arrangement, returns fee_id"""
        with self.conn.cursor() as cur:
            # Check for existing
            cur.execute(
                "SELECT fee_id FROM billing_fee_arrangement WHERE engagement_id = %s",
                (engagement_id,)
            )
            result = cur.fetchone()
            
            if result:
                fee_id = result[0]
                # Update
                cur.execute("""
                    UPDATE billing_fee_arrangement SET
                        raw_text = %s,
                        parsed_at = NOW(),
                        parser_version = 'kimi-v1',
                        updated_at = NOW()
                    WHERE fee_id = %s
                """, (data.fee_arrangement_text, fee_id))
            else:
                # Create new
                cur.execute("""
                    INSERT INTO billing_fee_arrangement (
                        engagement_id, raw_text, parser_version, parsed_at, created_at, updated_at
                    ) VALUES (%s, %s, 'kimi-v1', NOW(), NOW(), NOW())
                    RETURNING fee_id
                """, (engagement_id, data.fee_arrangement_text))
                fee_id = cur.fetchone()[0]
                self.stats['fee_arrangements_created'] += 1
            
            return fee_id
    
    def upsert_milestones(self, engagement_id: int, fee_id: int, milestones: List[Milestone]):
        """Upsert milestones"""
        with self.conn.cursor() as cur:
            for i, m in enumerate(milestones):
                # Check if exists
                cur.execute("""
                    SELECT milestone_id, completed, amount_value
                    FROM billing_milestone
                    WHERE engagement_id = %s AND ordinal = %s
                """, (engagement_id, m.ordinal))
                
                result = cur.fetchone()
                
                if result:
                    milestone_id, existing_completed, existing_amount = result
                    
                    # Update
                    cur.execute("""
                        UPDATE billing_milestone SET
                            description = %s,
                            amount_value = %s,
                            amount_currency = %s,
                            percent_value = %s,
                            is_percent = %s,
                            completed = %s,
                            completion_source = CASE WHEN %s THEN 'excel_strikethrough' ELSE completion_source END,
                            completion_date = CASE WHEN %s AND NOT COALESCE(completed, FALSE) THEN NOW()::date ELSE completion_date END,
                            updated_at = NOW(),
                            sort_order = %s
                        WHERE milestone_id = %s
                    """, (
                        m.description, m.amount, m.currency,
                        m.percent, m.percent is not None,
                        m.completed, m.completed, m.completed,
                        i, milestone_id
                    ))
                    self.stats['milestones_updated'] += 1
                else:
                    # Insert
                    cur.execute("""
                        INSERT INTO billing_milestone (
                            engagement_id, fee_id, ordinal, description,
                            amount_value, amount_currency, percent_value, is_percent,
                            completed, completion_source, completion_date,
                            created_at, updated_at, sort_order, raw_fragment
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW(), %s, %s)
                    """, (
                        engagement_id, fee_id, m.ordinal, m.description,
                        m.amount, m.currency, m.percent, m.percent is not None,
                        m.completed,
                        'excel_strikethrough' if m.completed else None,
                        datetime.now().date() if m.completed else None,
                        i,
                        m.description[:100]  # raw_fragment
                    ))
                    self.stats['milestones_inserted'] += 1
    
    def commit(self):
        if self.conn and not self.dry_run:
            self.conn.commit()
    
    def rollback(self):
        if self.conn:
            self.conn.rollback()


def parse_money(value) -> float:
    if value is None:
        return 0.0
    s = str(value).strip().replace(",", "").replace("$", "").replace("，", "")
    if s in ("", "-", "—", "–"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def load_excel_data(excel_path: str, specific_rows: List[int] = None) -> List[ProjectData]:
    print(f"📂 Loading Excel: {excel_path}")
    
    # Load data workbook
    wb_data = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    sheet_data = wb_data['Transactions']
    
    # Load format workbook (for strikethrough) - just once!
    wb_format = openpyxl.load_workbook(excel_path, data_only=False, read_only=True)
    sheet_format = wb_format['Transactions']
    
    strike_extractor = ExcelStrikethroughExtractor(excel_path)
    parser = KimiMilestoneParser()
    
    projects = []
    last_cm_no = ""  # Track last CM number for filling blanks
    
    if specific_rows:
        rows_to_process = [(min(specific_rows), max(specific_rows))]
    else:
        rows_to_process = [(5, sheet_data.max_row)]
    
    for min_row, max_row in rows_to_process:
        for idx, row in enumerate(sheet_data.iter_rows(min_row=min_row, max_row=max_row, values_only=True), start=min_row):
            if len(row) < 9:
                continue
            
            project_name = str(row[1]) if row[1] else ""
            client_name = str(row[2]) if row[2] else ""
            cm_no = str(row[3]) if row[3] else ""
            attorney = str(row[4]) if row[4] else ""
            sca = str(row[5]) if row[5] else ""
            
            # Handle missing CM numbers - use previous row's CM
            if not cm_no or cm_no.strip() == "" or cm_no.strip().lower() == "none":
                if last_cm_no:
                    cm_no = last_cm_no
                else:
                    continue
            else:
                last_cm_no = cm_no
            
            fees = parse_money(row[6])
            billing = parse_money(row[7])
            fee_text = str(row[8]) if row[8] else ""
            
            collection = parse_money(row[9] if len(row) > 9 else 0)
            billing_credit_usd = parse_money(row[10] if len(row) > 10 else 0)
            ubt_usd = parse_money(row[11] if len(row) > 11 else 0)
            ar_usd = parse_money(row[12] if len(row) > 12 else 0)
            billing_credit_cny = parse_money(row[13] if len(row) > 13 else 0)
            ubt_cny = parse_money(row[14] if len(row) > 14 else 0)
            
            finance_comment = str(row[15]) if len(row) > 15 and row[15] else ""
            remarks = str(row[16]) if len(row) > 16 and row[16] else ""
            matter_notes = str(row[17]) if len(row) > 17 and row[17] else ""
            
            # Get cell strikethrough - fast lookup from already-loaded workbook
            cell_strike = False
            try:
                cell = sheet_format.cell(idx, 9)  # Column I (Fee Arrangement)
                if cell.font:
                    cell_strike = cell.font.strike or False
            except Exception:
                pass
            
            struck_ordinals = strike_extractor.extract_for_cell(fee_text, cell_strike)
            milestones = parser.parse(fee_text, struck_ordinals)
            
            project = ProjectData(
                row_num=idx,
                project_name=project_name,
                client_name=client_name,
                cm_no=cm_no,
                attorney_in_charge=attorney,
                sca=sca,
                fees_usd=fees,
                billing_usd=billing,
                collection_usd=collection,
                billing_credit_usd=billing_credit_usd,
                ubt_usd=ubt_usd,
                ar_usd=ar_usd,
                billing_credit_cny=billing_credit_cny,
                ubt_cny=ubt_cny,
                finance_comment=finance_comment,
                remarks=remarks,
                matter_notes=matter_notes,
                fee_arrangement_text=fee_text,
                milestones=milestones,
                completed_milestones=struck_ordinals
            )
            
            projects.append(project)
    
    wb_data.close()
    wb_format.close()
    return projects


def main():
    parser = argparse.ArgumentParser(description='Update billing database with Kimi parser')
    parser.add_argument('--dry-run', action='store_true', help='Preview changes')
    parser.add_argument('--row', type=int, action='append', help='Process specific row(s)')
    parser.add_argument('--excel', default=EXCEL_FILE, help='Path to Excel file')
    parser.add_argument('--db-url', default=DATABASE_URL, help='Database URL')
    
    args = parser.parse_args()
    
    print("="*80)
    print("🤖 KIMI BILLING DATABASE UPDATER")
    print("="*80)
    print(f"Mode: {'DRY RUN' if args.dry_run else 'LIVE'}")
    print(f"Excel: {args.excel}")
    print("="*80)
    
    # Load data
    projects = load_excel_data(args.excel, args.row)
    print(f"\n📊 Loaded {len(projects)} projects")
    
    # Show sample
    print("\n📝 Sample data:")
    for p in projects[:3]:
        print(f"\n  Row {p.row_num}: CM={p.cm_no}")
        print(f"    Client: {p.client_name[:40]}...")
        print(f"    Financials: Billing=${p.billing_usd:,.0f}, Collected=${p.collection_usd:,.0f}")
        print(f"    Milestones: {len(p.milestones)}")
        for m in p.milestones[:2]:
            status = "✅" if m.completed else "⏳"
            amt = f"${m.amount:,.0f}" if m.amount else "TBD"
            print(f"      {status} {m.ordinal}: {amt}")
    
    if args.dry_run:
        print("\n" + "="*80)
        print("🏁 DRY RUN COMPLETE - No changes made")
        print("="*80)
        return
    
    # Apply changes
    print("\n" + "="*80)
    print("🔄 APPLYING CHANGES...")
    print("="*80)
    
    updater = DatabaseUpdater(args.db_url, dry_run=args.dry_run)
    
    try:
        updater.connect()
        
        for i, project in enumerate(projects, 1):
            print(f"\n[{i}/{len(projects)}] CM: {project.cm_no}")
            
            # Find project by CM number first, then by name
            project_id = updater.get_project_by_cm(project.cm_no)
            
            if not project_id:
                project_id = updater.get_project_by_name(project.project_name, project.client_name)
            
            if not project_id:
                print(f"    ⚠️ Project not found in DB: CM={project.cm_no}, Name={project.project_name[:30]}...")
                continue
            
            print(f"    Project ID: {project_id}")
            
            # Update project
            updater.update_project(project_id, project)
            
            # Update CM number/financials
            cm_id = updater.get_or_create_cm_no(project_id, project.cm_no, project)
            print(f"    CM ID: {cm_id}")
            
            # Get/create engagement
            engagement_id = updater.get_or_create_engagement(project_id, cm_id, project)
            print(f"    Engagement ID: {engagement_id}")
            
            # Get/create fee arrangement
            fee_id = updater.get_or_create_fee_arrangement(engagement_id, project)
            
            # Upsert milestones
            if project.milestones:
                updater.upsert_milestones(engagement_id, fee_id, project.milestones)
                completed = sum(1 for m in project.milestones if m.completed)
                print(f"    Milestones: {len(project.milestones)} total, {completed} completed")
        
        updater.commit()
        
        print("\n" + "="*80)
        print("✅ DATABASE UPDATE COMPLETE")
        print("="*80)
        print(f"Projects updated: {updater.stats['projects_updated']}")
        print(f"CM numbers updated: {updater.stats['cm_no_updated']}")
        print(f"Engagements created: {updater.stats['engagements_created']}")
        print(f"Fee arrangements created: {updater.stats['fee_arrangements_created']}")
        print(f"Milestones inserted: {updater.stats['milestones_inserted']}")
        print(f"Milestones updated: {updater.stats['milestones_updated']}")
        
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        updater.rollback()
        raise
    finally:
        updater.close()


if __name__ == "__main__":
    main()
