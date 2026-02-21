#!/usr/bin/env python3
"""
Smart Milestone Parser (Rule-Based)

Intelligently parses milestone text without requiring AI/LLM.
Handles various formats including:
- (a), (b), (c) or (1), (2), (3) ordinals
- Mixed Chinese/English text
- Various separators (-, —, etc.)
- Amounts with or without currency symbols
- Conditional milestones (no fixed amount)
- Strikethrough detection for completion status

Usage:
    python3 src/scripts/smart-parse-milestones.py [--dry-run] [--limit N]
"""

import os
import sys
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
from typing import Dict, List, Optional, Set, Tuple, Any
from dataclasses import dataclass, asdict

import openpyxl
import psycopg2
from psycopg2.extras import RealDictCursor

# Configuration
DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://postgres:qtSnaaSaelqHVydazTmViwejXbkkZxVY@crossover.proxy.rlwy.net:15782/railway",
)

DEFAULT_EXCEL_FILE = "/Users/timli/Library/CloudStorage/OneDrive-Personal/Coding/staffing-tracker/Billing/HKCM Project List (2026.02.12).xlsx"
EXCEL_FILE = os.environ.get("EXCEL_FILE", DEFAULT_EXCEL_FILE)

NS = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}


@dataclass
class Milestone:
    ordinal: str
    title: str
    description: str
    amount_usd: Optional[float]
    amount_cny: Optional[float]
    percentage: Optional[float]
    is_conditional: bool
    completed: bool
    raw_text: str


class SmartMilestoneParser:
    """Intelligently parses milestone text using heuristics"""
    
    # Pattern to find ordinal markers: (a), (b), (1), (2), etc.
    ORDINAL_PATTERN = re.compile(r'\(([a-z0-9])\)', re.IGNORECASE)
    
    # Pattern to find amounts: $95,000 or 95,000 or $95,000.00
    AMOUNT_PATTERN = re.compile(r'[\$¥]?\s*([\d,]+(?:\.\d{2})?)', re.IGNORECASE)
    
    # Pattern to find percentages: (50%) or 50%
    PERCENT_PATTERN = re.compile(r'\((\d+(?:\.\d+)?)%\)|(\d+(?:\.\d+)?)%')
    
    # Pattern to find LSD dates
    LSD_PATTERN = re.compile(r'\(LSD:\s*([^)]+)\)', re.IGNORECASE)
    
    # Keywords indicating conditional/adjustable amounts
    CONDITIONAL_KEYWORDS = [
        'calculated', 'adjust', '差额', 'difference', 'actual', '按小时',
        'hourly', 'depend', 'depending', 'based on', '根据', '按'
    ]
    
    def parse(self, text: str, strikethrough_ordinals: Set[str]) -> Tuple[List[Milestone], Dict]:
        """Parse fee arrangement text into milestones"""
        
        milestones = []
        
        # Extract LSD
        lsd_match = self.LSD_PATTERN.search(text)
        lsd_raw = lsd_match.group(1) if lsd_match else None
        lsd_date = self._parse_lsd_date(lsd_raw) if lsd_raw else None
        
        # Find all ordinal positions
        ordinals = list(self.ORDINAL_PATTERN.finditer(text))
        
        if not ordinals:
            return [], {'lsd_date': lsd_date, 'lsd_raw': lsd_raw}
        
        # Split text by ordinals
        for i, match in enumerate(ordinals):
            ordinal = f"({match.group(1).lower()})"
            start_pos = match.start()
            
            # Determine end position (next ordinal or end of text)
            if i < len(ordinals) - 1:
                end_pos = ordinals[i + 1].start()
            else:
                end_pos = len(text)
            
            milestone_text = text[start_pos:end_pos].strip()
            
            # Extract description (text after ordinal)
            desc_match = re.search(r'\([a-z0-9]\)\s*(.+?)(?:\(|$)', milestone_text, re.IGNORECASE | re.DOTALL)
            description = desc_match.group(1).strip() if desc_match else milestone_text
            
            # Clean up description
            description = description.replace('\n', ' ').strip()
            
            # Extract amount
            amount_usd = self._extract_amount(milestone_text)
            
            # Extract percentage
            percent_match = self.PERCENT_PATTERN.search(milestone_text)
            percentage = float(percent_match.group(1) or percent_match.group(2)) if percent_match else None
            
            # Check if conditional
            is_conditional = self._is_conditional(milestone_text)
            
            # Determine completion from strikethrough
            completed = ordinal in strikethrough_ordinals
            
            milestones.append(Milestone(
                ordinal=ordinal,
                title=description[:100],
                description=description,
                amount_usd=amount_usd if not is_conditional else None,
                amount_cny=None,
                percentage=percentage,
                is_conditional=is_conditional or amount_usd is None,
                completed=completed,
                raw_text=milestone_text
            ))
        
        # Extract bonus info
        bonus_usd, bonus_cny, bonus_desc = self._extract_bonus(text)
        
        return milestones, {
            'lsd_date': lsd_date,
            'lsd_raw': lsd_raw,
            'bonus_usd': bonus_usd,
            'bonus_cny': bonus_cny,
            'bonus_description': bonus_desc
        }
    
    def _parse_lsd_date(self, lsd_text: str) -> Optional[str]:
        """Parse LSD date string to YYYY-MM-DD"""
        if not lsd_text:
            return None
        
        # Try various formats
        patterns = [
            (r'(\d{1,2})\s+(\w{3,})\s+(\d{4})', self._parse_english_date),
            (r'(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日', self._parse_chinese_date),
        ]
        
        for pattern, parser in patterns:
            match = re.search(pattern, lsd_text, re.IGNORECASE)
            if match:
                return parser(match)
        
        return None
    
    def _parse_english_date(self, match) -> Optional[str]:
        day, month_str, year = match.groups()
        month_map = {
            'jan': 1, 'january': 1, 'feb': 2, 'february': 2, 'mar': 3, 'march': 3,
            'apr': 4, 'april': 4, 'may': 5, 'jun': 6, 'june': 6,
            'jul': 7, 'july': 7, 'aug': 8, 'august': 8, 'sep': 9, 'september': 9,
            'oct': 10, 'october': 10, 'nov': 11, 'november': 11, 'dec': 12, 'december': 12
        }
        month = month_map.get(month_str.lower())
        if month:
            return f"{year}-{month:02d}-{int(day):02d}"
        return None
    
    def _parse_chinese_date(self, match) -> str:
        year, month, day = match.groups()
        return f"{year}-{int(month):02d}-{int(day):02d}"
    
    def _extract_amount(self, text: str) -> Optional[float]:
        """Extract dollar amount from text"""
        # Look for patterns like - 95,000 or $95,000 or — 95,000
        patterns = [
            r'[-—–]\s*[\$¥]?\s*([\d,]+(?:\.\d{2})?)\s*$',  # End of line
            r'[-—–]\s*[\$¥]?\s*([\d,]+(?:\.\d{2})?)\s*\n',  # Before newline
            r'[-—–]\s*[\$¥]?\s*([\d,]+(?:\.\d{2})?)\s*(?:USD|CNY|$)',  # Before currency or end
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.MULTILINE)
            if match:
                try:
                    return float(match.group(1).replace(',', ''))
                except ValueError:
                    continue
        
        return None
    
    def _is_conditional(self, text: str) -> bool:
        """Check if milestone has conditional/adjustable amount"""
        text_lower = text.lower()
        for keyword in self.CONDITIONAL_KEYWORDS:
            if keyword.lower() in text_lower:
                return True
        return False
    
    def _extract_bonus(self, text: str) -> Tuple[Optional[float], Optional[float], Optional[str]]:
        """Extract bonus information"""
        bonus_usd = None
        bonus_cny = None
        bonus_desc = None
        
        # Chinese USD: 不低于X万美元奖金 or X万美元奖金
        match = re.search(r'(?:不低于)?(\d+(?:\.\d+)?)万美元奖金', text)
        if match:
            bonus_usd = float(match.group(1)) * 10000
            bonus_desc = match.group(0)
        
        # Chinese CNY: 不低于X万人民币奖金
        match = re.search(r'(?:不低于)?(\d+(?:\.\d+)?)万(?:人民币|元)奖金', text)
        if match:
            bonus_cny = float(match.group(1)) * 10000
            bonus_desc = match.group(0)
        
        # English: bonus of $X,XXX or bonus: $X,XXX
        match = re.search(r'bonus[:\s]+\$\s*([\d,]+)', text, re.IGNORECASE)
        if match:
            bonus_usd = float(match.group(1).replace(',', ''))
            bonus_desc = match.group(0)
        
        return bonus_usd, bonus_cny, bonus_desc


# Database updater (same as v2 script)
class DatabaseUpdater:
    def __init__(self, db_url: str):
        self.conn = psycopg2.connect(db_url)
        self.cur = self.conn.cursor(cursor_factory=RealDictCursor)
        self.stats = {'created': 0, 'updated': 0, 'errors': 0}
    
    def close(self):
        self.cur.close()
        self.conn.close()
    
    def update(self, project_name: str, cm_no: str, milestones: List[Milestone], 
               lsd_date: Optional[str], lsd_raw: Optional[str],
               bonus_usd: Optional[float], bonus_cny: Optional[float], 
               bonus_desc: Optional[str], raw_text: str):
        try:
            # Find C/M
            self.cur.execute(
                "SELECT cm_id FROM billing_project_cm_no WHERE cm_no = %s",
                (cm_no,)
            )
            cm = self.cur.fetchone()
            if not cm:
                print(f"  ⚠️ C/M not found: {cm_no}")
                return
            
            # Find engagement
            self.cur.execute(
                "SELECT engagement_id FROM billing_engagement WHERE cm_id = %s",
                (cm['cm_id'],)
            )
            eng = self.cur.fetchone()
            if not eng:
                print(f"  ⚠️ Engagement not found")
                return
            
            engagement_id = eng['engagement_id']
            
            # Get or create fee arrangement
            self.cur.execute(
                "SELECT fee_id FROM billing_fee_arrangement WHERE engagement_id = %s",
                (engagement_id,)
            )
            fee = self.cur.fetchone()
            
            if fee:
                fee_id = fee['fee_id']
                self.cur.execute("""
                    UPDATE billing_fee_arrangement 
                    SET raw_text = %s, lsd_date = %s, lsd_raw = %s,
                        bonus_description = %s, bonus_amount_usd = %s, bonus_amount_cny = %s
                    WHERE fee_id = %s
                """, (raw_text, lsd_date, lsd_raw, bonus_desc, bonus_usd, bonus_cny, fee_id))
            else:
                self.cur.execute("""
                    INSERT INTO billing_fee_arrangement 
                    (engagement_id, raw_text, lsd_date, lsd_raw, 
                     bonus_description, bonus_amount_usd, bonus_amount_cny, parsed_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
                    RETURNING fee_id
                """, (engagement_id, raw_text, lsd_date, lsd_raw, 
                      bonus_desc, bonus_usd, bonus_cny))
                fee_id = self.cur.fetchone()['fee_id']
            
            # Upsert milestones
            for m in milestones:
                self._upsert_milestone(fee_id, m)
            
            self.conn.commit()
            
        except Exception as e:
            self.conn.rollback()
            print(f"  ❌ Error: {e}")
            self.stats['errors'] += 1
    
    def _upsert_milestone(self, fee_id: int, m: Milestone):
        self.cur.execute(
            "SELECT milestone_id, completed FROM billing_milestone WHERE fee_id = %s AND ordinal = %s",
            (fee_id, m.ordinal)
        )
        existing = self.cur.fetchone()
        
        if existing:
            # Update
            self.cur.execute("""
                UPDATE billing_milestone 
                SET title = %s, description = %s, amount_value = %s,
                    is_percent = %s, percent_value = %s, updated_at = NOW()
                WHERE milestone_id = %s
            """, (m.title, m.description, m.amount_usd,
                  m.percentage is not None, m.percentage, existing['milestone_id']))
            
            if m.completed and not existing['completed']:
                self.cur.execute("""
                    UPDATE billing_milestone 
                    SET completed = true, completion_source = 'excel_strikethrough',
                        completion_date = CURRENT_DATE
                    WHERE milestone_id = %s
                """, (existing['milestone_id'],))
                self.stats['updated'] += 1
        else:
            # Create
            sort_order = ord(m.ordinal[1].lower()) - ord('a') + 1 if m.ordinal[1].isalpha() else int(m.ordinal[1])
            self.cur.execute("""
                INSERT INTO billing_milestone 
                (fee_id, ordinal, title, description, amount_value, amount_currency,
                 is_percent, percent_value, completed, completion_source, completion_date, 
                 raw_fragment, sort_order)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (fee_id, m.ordinal, m.title, m.description, m.amount_usd,
                  'USD' if m.amount_usd else None, m.percentage is not None, m.percentage,
                  m.completed, 'excel_strikethrough' if m.completed else None,
                  datetime.now().date() if m.completed else None, m.raw_text, sort_order))
            self.stats['created'] += 1


def extract_strikethrough(excel_path: str) -> Dict[int, Set[str]]:
    """Extract strikethrough ordinals from Excel XML"""
    data = {}
    try:
        with zipfile.ZipFile(excel_path, 'r') as zf:
            # Read shared strings
            ss_list = []
            if 'xl/sharedStrings.xml' in zf.namelist():
                root = ET.fromstring(zf.read('xl/sharedStrings.xml'))
                for si in root.findall('.//main:si', NS):
                    parts = []
                    has_strike = False
                    for r in si.findall('.//main:r', NS):
                        t = r.find('.//main:t', NS)
                        text = t.text if t is not None else ""
                        rpr = r.find('.//main:rPr', NS)
                        struck = False
                        if rpr is not None and rpr.find('.//main:strike', NS) is not None:
                            struck = True
                            has_strike = True
                        parts.append({'text': text, 'struck': struck})
                    ss_list.append({'parts': parts, 'has_strike': has_strike})
            
            # Read worksheet
            root = ET.fromstring(zf.read('xl/worksheets/sheet1.xml'))
            for row in root.findall('.//main:row', NS):
                row_num = int(row.get('r', 0))
                for cell in row.findall('.//main:c', NS):
                    if cell.get('r', '').startswith('I') and cell.get('t') == 's':
                        v = cell.find('.//main:v', NS)
                        if v is not None:
                            try:
                                ss = ss_list[int(v.text)]
                                if ss['has_strike']:
                                    ordinals = set()
                                    for part in ss['parts']:
                                        if part['struck']:
                                            for m in re.finditer(r'\(([a-z0-9])\)', part['text'], re.I):
                                                ordinals.add(f"({m[1].lower()})")
                                    if ordinals:
                                        data[row_num] = ordinals
                            except (ValueError, IndexError):
                                pass
    except Exception as e:
        print(f"Warning: {e}")
    return data


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run', action='store_true')
    parser.add_argument('--limit', type=int)
    parser.add_argument('--project', type=str)
    args = parser.parse_args()
    
    print("=" * 80)
    print("SMART MILESTONE PARSER (Rule-Based)")
    print("=" * 80)
    
    # Extract strikethrough
    print("\nExtracting strikethrough formatting...")
    strikethrough_data = extract_strikethrough(EXCEL_FILE)
    print(f"Found {len(strikethrough_data)} rows with strikethrough")
    
    # Load Excel
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    sheet = wb['Transactions']
    
    parser = SmartMilestoneParser()
    
    # Collect rows to process
    rows = []
    for row_idx in range(5, sheet.max_row + 1):
        project = sheet.cell(row_idx, 3).value
        if not project:
            continue
        if args.project and args.project.lower() not in str(project).lower():
            continue
        
        cm_no = str(sheet.cell(row_idx, 5).value or "").strip()
        fee_text = str(sheet.cell(row_idx, 9).value or "")
        
        if fee_text and len(fee_text) > 10:
            rows.append({
                'row': row_idx,
                'project': project,
                'cm_no': cm_no,
                'text': fee_text,
                'completed': strikethrough_data.get(row_idx, set())
            })
    
    wb.close()
    
    if args.limit:
        rows = rows[:args.limit]
    
    print(f"\nProcessing {len(rows)} fee arrangements...")
    
    # Process
    updater = None if args.dry_run else DatabaseUpdater(DATABASE_URL)
    
    try:
        for i, item in enumerate(rows, 1):
            print(f"\n[{i}] Row {item['row']}: {item['project']}")
            print(f"    C/M: {item['cm_no']}")
            print(f"    Strikethrough: {item['completed']}")
            
            milestones, meta = parser.parse(item['text'], item['completed'])
            
            if milestones:
                print(f"    ✓ Parsed {len(milestones)} milestones:")
                for m in milestones:
                    status = "✅" if m.completed else "⏳"
                    amt = f"${m.amount_usd:,.0f}" if m.amount_usd else f"{m.percentage}%" if m.percentage else "conditional"
                    print(f"       {status} {m.ordinal} {m.title[:45]}... ({amt})")
                
                if not args.dry_run:
                    updater.update(
                        item['project'], item['cm_no'], milestones,
                        meta['lsd_date'], meta['lsd_raw'],
                        meta['bonus_usd'], meta['bonus_cny'], meta['bonus_description'],
                        item['text']
                    )
            else:
                print(f"    ⚠️ No milestones found")
        
        if updater:
            print("\n" + "=" * 80)
            print(f"SUMMARY: Created {updater.stats['created']}, Updated {updater.stats['updated']}, Errors {updater.stats['errors']}")
    
    finally:
        if updater:
            updater.close()
    
    print("\n✅ Done!")


if __name__ == "__main__":
    main()
