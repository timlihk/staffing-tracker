#!/usr/bin/env python3
"""
FIXED Billing Database Updater with Kimi Parser

Fixes:
1. Matches Excel company names (column E) to DB client_name
2. Fixed amount parser - excludes years like 2025, 2026
3. Better strikethrough detection
"""

import os
import sys
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime

import openpyxl
import psycopg2

DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://postgres:qtSnaaSaelqHVydazTmViwejXbkkZxVY@crossover.proxy.rlwy.net:15782/railway",
)

EXCEL_FILE = "/Users/timli/Library/CloudStorage/OneDrive-Personal/Coding/staffing-tracker/Billing/HKCM Project List (2026.02.12).xlsx"
NS = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}


def parse_money(value):
    if value is None:
        return 0.0
    s = str(value).strip().replace(",", "").replace("$", "").replace("，", "")
    if s in ("", "-", "—", "–"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def extract_amount_fixed(text):
    """
    Fixed amount extraction - excludes years like 2025, 2026
    """
    if not text:
        return None
    
    # Clean text
    text_clean = text.replace(',', '').replace('，', '')
    
    # Pattern 1: $100,000 or 100,000 USD (before currency indicators)
    patterns = [
        (r'[US\$\$￥]\s*([\d]+(?:\.\d{2})?)', 'currency_prefix'),  # $100000 or $100000.00
        (r'([\d]+(?:\.\d{2})?)\s*(?:USD|美元|美金|CNY|RMB|人民币|元|HKD|港币)', 'currency_suffix'),
        (r'[-–—]\s*([\d]{5,}(?:\.\d{2})?)\s*$', 'dash_suffix'),  # -100000 at end
        (r'\b([\d]{5,}(?:\.\d{2})?)\s*[-–]', 'dash_prefix'),      # 100000- at start of amount section
    ]
    
    for pattern, ptype in patterns:
        match = re.search(pattern, text_clean)
        if match:
            amount_str = match.group(1)
            try:
                amount = float(amount_str)
                # Validate: must be reasonable amount (not year)
                if 1000 <= amount <= 100000000:  # Between 1K and 100M
                    # Additional check: not a year
                    if not (2020 <= amount <= 2030):  # Not a year
                        return amount
            except ValueError:
                continue
    
    # Pattern 2: Look for amount after specific keywords
    # 支付195,000 or 支付 195,000 or - 195,000
    keyword_patterns = [
        r'支付\s*([\d]+)',
        r'支付\s*USD\s*([\d]+)',
        r'payment\s*(?:of\s*)?\$?\s*([\d]+)',
    ]
    
    for pattern in keyword_patterns:
        match = re.search(pattern, text_clean, re.IGNORECASE)
        if match:
            try:
                amount = float(match.group(1))
                if 1000 <= amount <= 100000000 and not (2020 <= amount <= 2030):
                    return amount
            except ValueError:
                continue
    
    return None


def parse_milestones_fixed(text, struck_ordinals):
    """Improved milestone parser with fixed amount extraction"""
    if not text:
        return []
    
    milestones = []
    seen = set()
    
    # Find all ordinals
    for match in re.finditer(r'\(([a-z])\)', text, re.I):
        ordinal = f"({match.group(1).lower()})"
        if ordinal in seen:
            continue
        seen.add(ordinal)
        
        start = match.end()
        # Find segment end (next ordinal or end of text)
        next_match = None
        for m2 in re.finditer(r'\(([a-z])\)', text[start:], re.I):
            next_match = m2
            break
        
        if next_match:
            segment = text[start:start + next_match.start()].strip()
        else:
            segment = text[start:].strip()
        
        # Clean segment - stop at newline
        segment = re.sub(r'\n.*$', '', segment, flags=re.DOTALL)
        # Stop at next major milestone indicator
        segment = re.sub(r'\s+\([a-z]\).*$', '', segment, flags=re.DOTALL)
        
        # Extract amount with fixed parser
        amount = extract_amount_fixed(segment)
        
        # Extract percentage
        percent = None
        pct_match = re.search(r'\((\d+(?:\.\d+)?)%\)', segment)
        if pct_match:
            try:
                percent = float(pct_match.group(1))
            except:
                pass
        
        # Check completion
        completed = ordinal in struck_ordinals or "*" in struck_ordinals
        
        # Detect currency
        currency = 'USD'
        if any(c in segment.upper() for c in ['RMB', 'CNY', '人民币', '元']):
            currency = 'CNY'
        elif any(c in segment.upper() for c in ['HKD', '港币', '港元']):
            currency = 'HKD'
        
        milestones.append({
            'ordinal': ordinal,
            'description': segment[:200],
            'amount': amount,
            'currency': currency,
            'percent': percent,
            'completed': completed
        })
    
    return milestones


def main():
    print("="*80)
    print("🤖 KIMI BILLING DATABASE UPDATER - FIXED")
    print("="*80)
    
    # Load Excel
    print("\n1. Loading Excel...")
    wb_data = openpyxl.load_workbook(EXCEL_FILE, data_only=True, read_only=True)
    sheet_data = wb_data['Transactions']
    
    wb_format = openpyxl.load_workbook(EXCEL_FILE, data_only=False, read_only=True)
    sheet_format = wb_format['Transactions']
    print(f"   Sheet has {sheet_data.max_row} rows")
    
    # Load strikethrough data from XML
    print("\n2. Loading strikethrough data...")
    struck_data = {}
    with zipfile.ZipFile(EXCEL_FILE, 'r') as zf:
        ss_xml = zf.read('xl/sharedStrings.xml')
        root = ET.fromstring(ss_xml)
        
        for idx, si in enumerate(root.findall('.//main:si', NS)):
            parts = []
            has_strike = False
            for r in si.findall('.//main:r', NS):
                t = r.find('.//main:t', NS)
                text = t.text if t is not None else ""
                rpr = r.find('.//main:rPr', NS)
                struck = False
                if rpr is not None:
                    strike = rpr.find('.//main:strike', NS)
                    if strike is not None:
                        val = strike.get('val', 'true')
                        if val in ('true', '1', 'single'):
                            struck = True
                            has_strike = True
                parts.append({'text': text, 'struck': struck})
            
            full_text = ''.join(p['text'] for p in parts)
            struck_ordinals = set()
            for part in parts:
                if part['struck']:
                    for m in re.finditer(r'\(([a-z])\)', part['text'], re.I):
                        struck_ordinals.add(f"({m[1].lower()})")
            
            if struck_ordinals or has_strike:
                struck_data[idx] = {
                    'ordinals': struck_ordinals,
                    'has_strike': has_strike,
                    'text': full_text[:100]
                }
    
    print(f"   Found {len(struck_data)} cells with strikethrough")
    
    # Connect to database
    print("\n3. Connecting to database...")
    conn = psycopg2.connect(DATABASE_URL)
    conn.autocommit = False
    cur = conn.cursor()
    print("   Connected!")
    
    # Load CM number mapping
    print("\n4. Loading CM mapping...")
    cur.execute("SELECT cm_no, cm_id, project_id FROM billing_project_cm_no")
    cm_map = {}
    for row in cur.fetchall():
        cm_no, cm_id, project_id = row
        if cm_no:
            cm_map[cm_no.lower()] = {'cm_id': cm_id, 'project_id': project_id}
    
    print(f"   Cached {len(cm_map)} CM numbers")
    
    # Process rows
    print("\n5. Processing projects...")
    print("-"*80)
    
    stats = {
        'processed': 0,
        'matched_by_cm': 0,
        'new_projects': 0,
        'milestones': 0,
        'milestones_with_amount': 0,
        'completed_milestones': 0
    }
    
    last_company = ""
    
    for row_idx in range(5, sheet_data.max_row + 1):
        row = [cell.value for cell in sheet_data[row_idx]]
        
        if len(row) < 9:
            continue
        
        # Correct column mapping based on Excel structure:
        # Col B (index 1): No.
        # Col C (index 2): Project Name
        # Col D (index 3): Client Name
        # Col E (index 4): C/M No#
        # Col F (index 5): Attorney in Charge
        # Col G (index 6): SCA
        # Col H (index 7): Fees (US$)
        # Col I (index 8): (a) - Fee Arrangement
        
        project_name = str(row[2]) if row[2] else ""  # Column C
        client_name = str(row[3]) if row[3] else ""   # Column D
        cm_no = str(row[4]) if row[4] else ""         # Column E (C/M No#)
        attorney = str(row[5]) if row[5] else ""      # Column F
        sca = str(row[6]) if row[6] else ""           # Column G
        
        # CM number is in column E
        company = cm_no.strip() if cm_no else ""
        
        # Handle empty CM - use previous row's CM
        if not company or company.lower() == "none":
            if last_company:
                company = last_company
            else:
                continue
        else:
            last_company = company
        
        fee_text = str(row[8]) if row[8] else ""  # Column I - Fee Arrangement
        if not fee_text or len(fee_text) < 10:
            continue
        
        # Get strikethrough
        cell_strike = False
        try:
            cell = sheet_format.cell(row_idx, 9)
            if cell.font:
                cell_strike = cell.font.strike or False
        except:
            pass
        
        # Build struck ordinals
        struck_ordinals = set()
        if cell_strike:
            struck_ordinals.add("*")
        
        # Check XML strikethrough
        for ss_idx, ss_info in struck_data.items():
            if company[:30].lower() in ss_info['text'].lower() or ss_info['text'][:30].lower() in fee_text.lower():
                struck_ordinals.update(ss_info['ordinals'])
        
        # Parse milestones
        milestones = parse_milestones_fixed(fee_text, struck_ordinals)
        
        # Find project by CM number
        project_info = None
        match_type = None
        
        # Try exact CM match
        if company.lower() in cm_map:
            project_info = cm_map[company.lower()]
            match_type = "existing"
            stats['matched_by_cm'] += 1
        else:
            # Create new project for this CM
            cur.execute("""
                INSERT INTO billing_project (project_name, client_name, base_currency, created_at, updated_at)
                VALUES (%s, %s, 'USD', NOW(), NOW())
                RETURNING project_id
            """, (project_name or company, client_name))
            
            project_id = cur.fetchone()[0]
            
            # Create CM number entry
            cur.execute("""
                INSERT INTO billing_project_cm_no (project_id, cm_no, is_primary, status)
                VALUES (%s, %s, FALSE, 'active')
                RETURNING cm_id
            """, (project_id, company))
            
            cm_id = cur.fetchone()[0]
            
            # Add to cache
            cm_map[company.lower()] = {'cm_id': cm_id, 'project_id': project_id}
            
            project_info = {'cm_id': cm_id, 'project_id': project_id}
            match_type = "new"
            stats['new_projects'] += 1
        
        stats['processed'] += 1
        stats['milestones'] += len(milestones)
        stats['milestones_with_amount'] += sum(1 for m in milestones if m['amount'])
        stats['completed_milestones'] += sum(1 for m in milestones if m['completed'])
        
        # Show progress
        if stats['processed'] % 20 == 0:
            print(f"   Progress: {stats['processed']} processed, "
                  f"{stats['matched_by_cm']} existing, "
                  f"{stats['new_projects']} new, "
                  f"{stats['milestones']} milestones")
        
        project_id = project_info['project_id']
        cm_id = project_info['cm_id']
        
        # Update project
        cur.execute("""
            UPDATE billing_project SET
                client_name = %s,
                attorney_in_charge = %s,
                sca = %s,
                updated_at = NOW()
            WHERE project_id = %s
        """, (client_name, attorney, sca, project_id))
        
        # Get or create engagement
        cur.execute("SELECT engagement_id FROM billing_engagement WHERE cm_id = %s", (cm_id,))
        result = cur.fetchone()
        
        if result:
            engagement_id = result[0]
        else:
            cur.execute("""
                INSERT INTO billing_engagement (project_id, cm_id, engagement_code, engagement_title, updated_at)
                VALUES (%s, %s, 'original', %s, NOW())
                RETURNING engagement_id
            """, (project_id, cm_id, project_name[:100]))
            engagement_id = cur.fetchone()[0]
        
        # Get or create fee arrangement
        cur.execute("SELECT fee_id FROM billing_fee_arrangement WHERE engagement_id = %s", (engagement_id,))
        result = cur.fetchone()
        
        if result:
            fee_id = result[0]
            cur.execute("""
                UPDATE billing_fee_arrangement 
                SET raw_text = %s, parser_version = 'kimi-v2', parsed_at = NOW(), updated_at = NOW()
                WHERE fee_id = %s
            """, (fee_text, fee_id))
        else:
            cur.execute("""
                INSERT INTO billing_fee_arrangement (engagement_id, raw_text, parser_version, parsed_at, created_at, updated_at)
                VALUES (%s, %s, 'kimi-v2', NOW(), NOW(), NOW())
                RETURNING fee_id
            """, (engagement_id, fee_text))
            fee_id = cur.fetchone()[0]
        
        # Upsert milestones
        for i, m in enumerate(milestones):
            cur.execute("""
                SELECT milestone_id FROM billing_milestone
                WHERE engagement_id = %s AND ordinal = %s
            """, (engagement_id, m['ordinal']))
            result = cur.fetchone()
            
            if result:
                cur.execute("""
                    UPDATE billing_milestone SET
                        description = %s,
                        amount_value = %s,
                        amount_currency = %s,
                        percent_value = %s,
                        is_percent = %s,
                        completed = %s,
                        completion_source = CASE WHEN %s THEN 'excel_strikethrough' ELSE completion_source END,
                        completion_date = CASE WHEN %s AND NOT COALESCE(completed, FALSE) THEN NOW()::date ELSE completion_date END,
                        updated_at = NOW(),
                        sort_order = %s,
                        raw_fragment = %s
                    WHERE milestone_id = %s
                """, (
                    m['description'], m['amount'], m['currency'], m['percent'],
                    m['percent'] is not None, m['completed'], m['completed'],
                    m['completed'], i, m['description'][:100], result[0]
                ))
            else:
                cur.execute("""
                    INSERT INTO billing_milestone (
                        engagement_id, fee_id, ordinal, description,
                        amount_value, amount_currency, percent_value, is_percent,
                        completed, completion_source, completion_date,
                        created_at, updated_at, sort_order, raw_fragment
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW(), %s, %s)
                """, (
                    engagement_id, fee_id, m['ordinal'], m['description'],
                    m['amount'], m['currency'], m['percent'], m['percent'] is not None,
                    m['completed'],
                    'excel_strikethrough' if m['completed'] else None,
                    datetime.now().date() if m['completed'] else None,
                    i, m['description'][:100]
                ))
        
        # Commit every 10 projects
        if stats['processed'] % 10 == 0:
            conn.commit()
    
    # Final commit
    conn.commit()
    
    print("-"*80)
    print("\n✅ COMPLETE!")
    print(f"   Projects processed: {stats['processed']}")
    print(f"   Existing projects: {stats['matched_by_cm']}")
    print(f"   New projects created: {stats['new_projects']}")
    print(f"   Total milestones: {stats['milestones']}")
    print(f"   Milestones with amount: {stats['milestones_with_amount']}")
    print(f"   Completed milestones: {stats['completed_milestones']}")
    
    cur.close()
    conn.close()
    wb_data.close()
    wb_format.close()


if __name__ == "__main__":
    main()
