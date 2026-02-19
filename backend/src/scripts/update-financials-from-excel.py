#!/usr/bin/env python3
"""
Update billing financial fields from HKCM Excel workbook.

Reads:
- "Time & Fees reports" sheet:
  - Mttr# (C/M) in column 10
  - Fees Billed in column 17
  - Collected in column 23
- "UA Report" sheet:
  - Client matter (C/M) in column 1
  - UBT in column 31

Updates table:
- billing_project_cm_no (billing_to_date_usd, collected_to_date_usd, ubt_usd, financials_updated_at)
"""

import os
from datetime import datetime

import openpyxl
import psycopg2


DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://postgres:qtSnaaSaelqHVydazTmViwejXbkkZxVY@crossover.proxy.rlwy.net:15782/railway",
)
EXCEL_FILE = os.environ.get("EXCEL_FILE")


def parse_number(value):
    if value is None:
        return 0.0
    s = str(value).strip().replace(",", "")
    if s == "":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def load_workbook_data(path):
    wb = openpyxl.load_workbook(path, data_only=True, keep_vba=False, read_only=True)

    time_fees = {}
    tf_sheet = wb["Time & Fees reports"] if "Time & Fees reports" in wb.sheetnames else None
    if tf_sheet is not None:
        for row in tf_sheet.iter_rows(min_row=1, values_only=True):
            cm_no = str(row[9]).strip() if len(row) > 9 and row[9] is not None else ""
            if not cm_no:
                continue
            fees_billed = parse_number(row[16]) if len(row) > 16 else 0.0
            collected = parse_number(row[22]) if len(row) > 22 else 0.0
            ar_value = parse_number(row[23]) if len(row) > 23 else 0.0
            time_fees[cm_no] = {
                "fees_billed": fees_billed,
                "collected": collected,
                "ar": ar_value,
            }

    ua = {}
    ua_sheet = wb["UA Report"] if "UA Report" in wb.sheetnames else None
    if ua_sheet is not None:
        for row in ua_sheet.iter_rows(min_row=2, values_only=True):
            cm_no = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
            if not cm_no:
                continue
            ubt = parse_number(row[30]) if len(row) > 30 else 0.0
            ua[cm_no] = {"ubt": ubt}

    # Fallback for consolidated HKCM files with only "Transactions" sheet.
    # Columns in row 4 headers:
    # E: C/M No, J: Billing (US$), K: Collection (US$), M: UBT (US$)
    if (not time_fees or not ua) and "Transactions" in wb.sheetnames:
        tx_sheet = wb["Transactions"]
        for row in tx_sheet.iter_rows(min_row=5, values_only=True):
            cm_no = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
            if not cm_no:
                continue
            billing = parse_number(row[9]) if len(row) > 9 else 0.0
            collection = parse_number(row[10]) if len(row) > 10 else 0.0
            ubt = parse_number(row[12]) if len(row) > 12 else 0.0
            time_fees[cm_no] = {
                "fees_billed": billing,
                "collected": collection,
                "ar": 0.0,
            }
            ua[cm_no] = {"ubt": ubt}

    wb.close()
    return time_fees, ua


def update_db(time_fees, ua):
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()

    cur.execute("SELECT cm_id, cm_no FROM billing_project_cm_no")
    cm_rows = cur.fetchall()

    matched = 0
    updated = 0

    for cm_id, cm_no in cm_rows:
        key = (cm_no or "").strip()
        tf = time_fees.get(key)
        ua_row = ua.get(key)
        if tf is None and ua_row is None:
            continue

        matched += 1

        set_parts = ["financials_updated_at = %s"]
        params = [datetime.utcnow()]

        if tf is not None and tf["fees_billed"] != 0:
            set_parts.append("billing_to_date_usd = %s")
            params.append(tf["fees_billed"])
        if tf is not None and tf["collected"] != 0:
            set_parts.append("collected_to_date_usd = %s")
            params.append(tf["collected"])
        if ua_row is not None and ua_row["ubt"] != 0:
            set_parts.append("ubt_usd = %s")
            params.append(ua_row["ubt"])

        if len(set_parts) == 1:
            continue

        params.append(cm_id)
        cur.execute(
            f"UPDATE billing_project_cm_no SET {', '.join(set_parts)} WHERE cm_id = %s",
            params,
        )
        updated += cur.rowcount

    conn.commit()
    cur.close()
    conn.close()
    return matched, updated


def main():
    if not EXCEL_FILE:
        raise RuntimeError("Set EXCEL_FILE environment variable to the workbook path")

    print(f"Loading workbook: {EXCEL_FILE}")
    time_fees, ua = load_workbook_data(EXCEL_FILE)
    print(f"Loaded Time & Fees rows: {len(time_fees)}")
    print(f"Loaded UA rows: {len(ua)}")

    matched, updated = update_db(time_fees, ua)
    print(f"Matched DB C/M rows: {matched}")
    print(f"Updated DB C/M rows: {updated}")


if __name__ == "__main__":
    main()
