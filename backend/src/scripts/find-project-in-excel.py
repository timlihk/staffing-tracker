#!/usr/bin/env python3
"""
Find which worksheet contains the project data
"""

import openpyxl

EXCEL_FILE = '/home/timlihk/staffing-tracker/billing-matter/HKCM Project List(81764217.1)_6Oct25.xlsx'
SEARCH_TERM = 'Salus'

print(f"Searching for '{SEARCH_TERM}' in all worksheets...")
print("Loading Excel file (fast mode)...\n")

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True, read_only=True)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    found = False

    # Search first 50 rows only
    max_search_rows = min(50, sheet.max_row)

    for row_idx in range(1, max_search_rows + 1):
        for cell in sheet[row_idx]:
            if cell.value and SEARCH_TERM.lower() in str(cell.value).lower():
                if not found:
                    print(f"✅ Found in worksheet: {sheet_name}")
                    print(f"   Row {row_idx}: {str(cell.value)[:100]}")
                    print(f"   Column: {cell.column_letter}")

                    # Print header row if we're past row 1
                    if row_idx > 1:
                        print(f"\n   Headers (row 1):")
                        for header_cell in sheet[1]:
                            if header_cell.value:
                                print(f"     Col {header_cell.column_letter}: {str(header_cell.value)[:40]}")
                    print()
                    found = True
                    break
        if found:
            break

wb.close()

print("Search complete!")
