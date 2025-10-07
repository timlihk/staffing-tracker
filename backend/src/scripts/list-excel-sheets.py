#!/usr/bin/env python3
"""
List all worksheets and their headers in the Excel file
"""

import openpyxl

EXCEL_FILE = '/home/timlihk/staffing-tracker/billing-matter/HKCM Project List(81764217.1)_6Oct25.xlsx'

print("Loading Excel file (fast mode)...")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True, read_only=True)

print(f"\nFound {len(wb.sheetnames)} worksheets:\n")

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"📋 {sheet_name}")
    print(f"   Rows: {sheet.max_row}, Columns: {sheet.max_column}")

    # Get first row headers
    headers = []
    for cell in sheet[1]:
        if cell.value:
            headers.append(str(cell.value)[:30])  # Limit to 30 chars

    if headers:
        print(f"   Headers: {', '.join(headers[:10])}")  # Show first 10 headers
        if len(headers) > 10:
            print(f"            ... and {len(headers) - 10} more")

    print()

wb.close()
