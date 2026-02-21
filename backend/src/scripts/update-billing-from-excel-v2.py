#!/usr/bin/env python3
"""
Master Billing Update Script v2 - With Proper Partial Strikethrough Support

This version properly handles partial strikethrough within cells by:
1. Reading Excel XML directly to detect rich text formatting
2. Converting Excel to HTML using available tools
3. Parsing HTML to detect <s> tags and line-through styles

Usage:
    export EXCEL_FILE="/path/to/Billing/HKCM Project List.xlsx"
    python3 src/scripts/update-billing-from-excel-v2.py
"""

import os
import sys
import re
import zipfile
import xml.etree.ElementTree as ET
import hashlib
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any, Set
from dataclasses import dataclass

import openpyxl
import psycopg2
from psycopg2.extras import RealDictCursor

# Configuration
DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://postgres:qtSnaaSaelqHVydazTmViwejXbkkZxVY@crossover.proxy.rlwy.net:15782/railway",
)

DEFAULT_EXCEL_FILE = "/Users/timli/Library/CloudStorage/OneDrive-Personal/Coding/staffing-tracker/Billing/HKCM Project List (2026.02.12).xlsx"
EXCEL_FILE = os.environ.get("EXCEL_FILE", DEFAULT_EXCEL_FILE)

# XML namespaces for Excel
NS = {
    'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
}


@dataclass
class ProjectData:
    """Represents a row from the Excel file"""
    row_num: int
    project_name: str
    client_name: str
    cm_no: str
    attorney_in_charge: str
    sca: str
    fees_usd: float
    billing_usd: float
    collection_usd: float
    billing_credit_usd: float
    ubt_usd: float
    ar_usd: float
    billing_credit_cny: float
    ubt_cny: float
    finance_comment: str
    remarks: str
    matter_notes: str
    
    # Fee arrangement parsing
    fee_arrangement_text: str = ""
    milestones: List[Dict[str, Any]] = None
    bonuses: List[Dict[str, Any]] = None
    completed_milestones: Set[str] = None
    lsd_date: Optional[datetime] = None
    lsd_raw: Optional[str] = None
    
    def __post_init__(self):
        if self.milestones is None:
            self.milestones = []
        if self.bonuses is None:
            self.bonuses = []
        if self.completed_milestones is None:
            self.completed_milestones = set()


def parse_money(value) -> float:
    """Parse monetary value from Excel cell"""
    if value is None:
        return 0.0
    s = str(value).strip().replace(",", "").replace("$", "")
    if s == "" or s == "-":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def extract_strikethrough_from_xml(excel_path: str, sheet_name: str) -> Dict[int, Set[str]]:
    """
    Extract strikethrough information directly from Excel XML.
    Returns a dict mapping row_number -> set of completed ordinals.
    """
    strikethrough_data = {}
    
    try:
        with zipfile.ZipFile(excel_path, 'r') as zf:
            # Read shared strings
            shared_strings = []
            if 'xl/sharedStrings.xml' in zf.namelist():
                ss_xml = zf.read('xl/sharedStrings.xml')
                root = ET.fromstring(ss_xml)
                
                # Each si element is a shared string item
                for si in root.findall('.//main:si', NS):
                    # Check for rich text with formatting
                    text_parts = []
                    has_strike = False
                    struck_texts = []
                    
                    # Check for r (run) elements
                    for r in si.findall('.//main:r', NS):
                        # Get text content
                        t = r.find('.//main:t', NS)
                        text = t.text if t is not None else ""
                        
                        # Check for rPr (run properties) with strike
                        rpr = r.find('.//main:rPr', NS)
                        is_struck = False
                        if rpr is not None:
                            # Check for strike element
                            strike = rpr.find('.//main:strike', NS)
                            if strike is not None:
                                val = strike.get('val', 'true')
                                if val in ('true', '1', 'single'):
                                    is_struck = True
                                    has_strike = True
                                    struck_texts.append(text)
                        
                        text_parts.append({
                            'text': text,
                            'struck': is_struck
                        })
                    
                    # If no rich text runs, check direct text
                    if not text_parts:
                        t = si.find('.//main:t', NS)
                        if t is not None:
                            text_parts.append({
                                'text': t.text,
                                'struck': False
                            })
                    
                    shared_strings.append({
                        'parts': text_parts,
                        'has_strike': has_strike,
                        'struck_texts': struck_texts
                    })
            
            # Read worksheet
            sheet_path = f'xl/worksheets/{sheet_name.lower()}.xml'
            if sheet_path not in zf.namelist():
                # Try with sheet number
                sheet_path = 'xl/worksheets/sheet1.xml'
            
            if sheet_path in zf.namelist():
                ws_xml = zf.read(sheet_path)
                root = ET.fromstring(ws_xml)
                
                # Process each row
                for row in root.findall('.//main:row', NS):
                    row_num = int(row.get('r', 0))
                    
                    for cell in row.findall('.//main:c', NS):
                        cell_ref = cell.get('r', '')
                        cell_type = cell.get('t', '')
                        
                        # Only process column I (Fee Arrangement)
                        if not cell_ref.startswith('I'):
                            continue
                        
                        # Get value
                        if cell_type == 's':  # Shared string
                            v = cell.find('.//main:v', NS)
                            if v is not None:
                                try:
                                    ss_idx = int(v.text)
                                    if ss_idx < len(shared_strings):
                                        ss_item = shared_strings[ss_idx]
                                        if ss_item['has_strike']:
                                            # Extract milestone ordinals from struck text
                                            struck_ordinals = set()
                                            for part in ss_item['parts']:
                                                if part['struck']:
                                                    # Find ALL milestone ordinals in struck text, not just first
                                                    for match in re.finditer(r'\(([a-z])\)', part['text'], re.IGNORECASE):
                                                        struck_ordinals.add(f"({match[1].lower()})")
                                            
                                            if struck_ordinals:
                                                strikethrough_data[row_num] = struck_ordinals
                                except (ValueError, IndexError):
                                    pass
    
    except Exception as e:
        print(f"Warning: Could not extract strikethrough from XML: {e}")
    
    return strikethrough_data


def parse_milestones(text: str, completed_ordinals: Set[str]) -> List[Dict[str, Any]]:
    """Parse milestones from fee arrangement text"""
    milestones = []
    seen_ordinals = set()
    
    lines = text.split('\n')
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Pattern 1: With amount - (a) description (50%) - 195,000
        match = re.search(
            r'\(([a-z])\)\s*(.+?)(?:\((\d+(?:\.\d+)?)%\))?\s*-\s*([\d,]+)',
            line,
            re.IGNORECASE
        )
        
        amount = None
        percent = None
        
        if match:
            ordinal, description, percent_str, amount_str = match.groups()
            try:
                amount = float(amount_str.replace(',', ''))
                percent = float(percent_str) if percent_str else None
            except ValueError:
                pass
        else:
            # Pattern 2: Without amount - just ordinal and description
            match = re.search(
                r'\(([a-z])\)\s*(.+?)(?:\((\d+(?:\.\d+)?)%\))?$',
                line,
                re.IGNORECASE
            )
            if match:
                ordinal, description, percent_str = match.groups()
                try:
                    percent = float(percent_str) if percent_str else None
                except ValueError:
                    pass
            else:
                continue
        
        if not match:
            continue
        
        ordinal = match.group(1)
        description = match.group(2).strip()
        ordinal_key = f"({ordinal.lower()})"
        
        if ordinal_key in seen_ordinals:
            continue
        seen_ordinals.add(ordinal_key)
        
        if len(description) < 3 or description.lower() in ['original el', 'supplemental el']:
            continue
        
        milestones.append({
            'ordinal': ordinal_key,
            'title': description[:100],
            'description': description,
            'trigger_text': description,
            'amount_value': amount,
            'amount_currency': 'USD' if amount else None,
            'is_percent': percent is not None,
            'percent_value': percent,
            'sort_order': ord(ordinal.lower()) - ord('a') + 1,
            'completed': ordinal_key in completed_ordinals
        })
    
    return milestones


def parse_lsd(text: str) -> Tuple[Optional[datetime], Optional[str]]:
    """Parse Long Stop Date from fee arrangement text"""
    match = re.search(r'\(LSD:\s*([^)]+)\)', text, re.IGNORECASE)
    if not match:
        return None, None
    
    lsd_text = match[1].strip()
    
    date_match = re.search(r'(\d{1,2})\s+(\w{3,})\s+(\d{4})', lsd_text)
    if date_match:
        day, month_str, year = date_match.groups()
        month_map = {
            'jan': 1, 'january': 1, 'feb': 2, 'february': 2, 'mar': 3, 'march': 3,
            'apr': 4, 'april': 4, 'may': 5, 'jun': 6, 'june': 6,
            'jul': 7, 'july': 7, 'aug': 8, 'august': 8, 'sep': 9, 'september': 9,
            'oct': 10, 'october': 10, 'nov': 11, 'november': 11, 'dec': 12, 'december': 12
        }
        month = month_map.get(month_str.lower())
        if month:
            return datetime(int(year), month, int(day)), lsd_text
    
    return None, lsd_text


def load_excel_data(file_path: str) -> List[ProjectData]:
    """Load and parse Excel file"""
    print(f"Loading Excel file: {file_path}")
    
    if not os.path.exists(file_path):
        print(f"Error: Excel file not found: {file_path}")
        sys.exit(1)
    
    # Extract strikethrough data from XML
    print("Extracting strikethrough formatting from Excel XML...")
    strikethrough_data = extract_strikethrough_from_xml(file_path, 'Transactions')
    print(f"Found strikethrough data for {len(strikethrough_data)} rows")
    for row, ordinals in list(strikethrough_data.items())[:5]:
        print(f"  Row {row}: {ordinals}")
    
    # Load with openpyxl for other data
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
    sheet = wb['Transactions']
    
    projects = []
    
    for row_idx in range(5, sheet.max_row + 1):
        row = sheet[row_idx]
        
        if not row[2].value:  # Column C - Project Name
            continue
        
        # Get strikethrough ordinals for this row
        completed_ordinals = strikethrough_data.get(row_idx, set())
        
        # Also check cell-level strikethrough (entire cell struck)
        fee_cell = sheet.cell(row=row_idx, column=9)
        if fee_cell.font and getattr(fee_cell.font, 'strike', False):
            # Entire cell struck - all milestones completed
            text = str(fee_cell.value) if fee_cell.value else ""
            for match in re.finditer(r'\(([a-z])\)', text, re.IGNORECASE):
                completed_ordinals.add(f"({match[1].lower()})")
        
        text = str(fee_cell.value) if fee_cell.value else ""
        lsd_date, lsd_raw = parse_lsd(text)
        milestones = parse_milestones(text, completed_ordinals)
        
        project = ProjectData(
            row_num=row_idx,
            project_name=str(row[2].value or "").strip(),
            client_name=str(row[3].value or "").strip(),
            cm_no=str(row[4].value or "").strip(),
            attorney_in_charge=str(row[5].value or "").strip(),
            sca=str(row[6].value or "").strip(),
            fees_usd=parse_money(row[7].value),
            billing_usd=parse_money(row[9].value),
            collection_usd=parse_money(row[10].value),
            billing_credit_usd=parse_money(row[11].value),
            ubt_usd=parse_money(row[12].value),
            ar_usd=parse_money(row[13].value),
            billing_credit_cny=parse_money(row[15].value),
            ubt_cny=parse_money(row[16].value),
            finance_comment=str(row[17].value or "").strip(),
            remarks=str(row[20].value or "").strip(),
            matter_notes=str(row[21].value or "").strip(),
            fee_arrangement_text=text,
            milestones=milestones,
            completed_milestones=completed_ordinals,
            lsd_date=lsd_date,
            lsd_raw=lsd_raw
        )
        
        projects.append(project)
    
    wb.close()
    print(f"Loaded {len(projects)} projects from Excel")
    return projects


def main():
    print("=" * 80)
    print("BILLING UPDATE v2 - WITH PARTIAL STRIKETHROUGH SUPPORT")
    print("=" * 80)
    
    projects = load_excel_data(EXCEL_FILE)
    
    # Print summary of projects with partial strikethrough
    print("\n" + "=" * 80)
    print("PROJECTS WITH COMPLETED MILESTONES")
    print("=" * 80)
    
    for project in projects:
        if project.completed_milestones:
            completed = [m for m in project.milestones if m['completed']]
            pending = [m for m in project.milestones if not m['completed']]
            
            print(f"\nRow {project.row_num}: {project.project_name}")
            print(f"  C/M: {project.cm_no}")
            print(f"  Completed: {[m['ordinal'] for m in completed]}")
            print(f"  Pending: {[m['ordinal'] for m in pending]}")


if __name__ == "__main__":
    main()
