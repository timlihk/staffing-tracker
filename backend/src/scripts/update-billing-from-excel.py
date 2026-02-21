#!/usr/bin/env python3
"""
Master Billing Update Script

Updates billing database from HKCM Project List Excel file.
Combines multiple updates:
1. Financial data (billing, collection, UBT, billing credit)
2. Fee arrangements with milestones
3. Strikethrough detection for completed milestones
4. Bonus parsing

Usage:
    cd backend
    export DATABASE_URL="your-db-url"
    export EXCEL_FILE="/path/to/Billing/HKCM Project List (2026.02.12).xlsx"
    python3 src/scripts/update-billing-from-excel.py
"""

import os
import sys
import re
import hashlib
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass

import openpyxl
from openpyxl.cell.cell import Cell
import psycopg2
from psycopg2.extras import RealDictCursor

# Configuration
DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://postgres:qtSnaaSaelqHVydazTmViwejXbkkZxVY@crossover.proxy.rlwy.net:15782/railway",
)

DEFAULT_EXCEL_FILE = "/Users/timli/Library/CloudStorage/OneDrive-Personal/Coding/staffing-tracker/Billing/HKCM Project List (2026.02.12).xlsx"
EXCEL_FILE = os.environ.get("EXCEL_FILE", DEFAULT_EXCEL_FILE)

# Color codes for terminal output
class Colors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'


def print_header(text: str):
    print(f"\n{Colors.HEADER}{'='*80}{Colors.ENDC}")
    print(f"{Colors.HEADER}{text.center(80)}{Colors.ENDC}")
    print(f"{Colors.HEADER}{'='*80}{Colors.ENDC}\n")


def print_success(text: str):
    print(f"{Colors.OKGREEN}✓ {text}{Colors.ENDC}")


def print_warning(text: str):
    print(f"{Colors.WARNING}⚠ {text}{Colors.ENDC}")


def print_error(text: str):
    print(f"{Colors.FAIL}✗ {text}{Colors.ENDC}")


def print_info(text: str):
    print(f"{Colors.OKBLUE}ℹ {text}{Colors.ENDC}")


@dataclass
class ProjectData:
    """Represents a row from the Excel file"""
    row_num: int
    project_name: str
    client_name: str
    cm_no: str
    attorney_in_charge: str
    sca: str
    fees_usd: float
    billing_usd: float
    collection_usd: float
    billing_credit_usd: float
    ubt_usd: float
    ar_usd: float
    billing_credit_cny: float
    ubt_cny: float
    finance_comment: str
    remarks: str
    matter_notes: str
    
    # Fee arrangement parsing
    fee_arrangement_text: str = ""
    milestones: List[Dict[str, Any]] = None
    bonuses: List[Dict[str, Any]] = None
    completed_milestones: set = None
    
    def __post_init__(self):
        if self.milestones is None:
            self.milestones = []
        if self.bonuses is None:
            self.bonuses = []
        if self.completed_milestones is None:
            self.completed_milestones = set()


def parse_money(value) -> float:
    """Parse monetary value from Excel cell"""
    if value is None:
        return 0.0
    s = str(value).strip().replace(",", "").replace("$", "")
    if s == "" or s == "-":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def extract_fee_arrangement_text(cell: Cell) -> Tuple[str, set]:
    """Extract fee arrangement text and detect strikethrough milestones"""
    if not cell or cell.value is None:
        return "", set()
    
    text = str(cell.value)
    completed = set()
    
    # Check for rich text with strikethrough
    if hasattr(cell.value, 'richText') and cell.value.richText:
        for run in cell.value.richText:
            if hasattr(run, 'font') and run.font and getattr(run.font, 'strike', False):
                # Extract milestone ordinal from struck text
                match = re.search(r'\(([a-z])\)', str(run.text), re.IGNORECASE)
                if match:
                    completed.add(f"({match[1].lower()})")
    
    # Check if entire cell has strikethrough
    if cell.font and getattr(cell.font, 'strike', False):
        matches = re.finditer(r'\(([a-z])\)', text, re.IGNORECASE)
        for match in matches:
            completed.add(f"({match[1].lower()})")
    
    return text, completed


def parse_lsd(text: str) -> Tuple[Optional[datetime], Optional[str]]:
    """Parse Long Stop Date from fee arrangement text"""
    match = re.search(r'\(LSD:\s*([^)]+)\)', text, re.IGNORECASE)
    if not match:
        return None, None
    
    lsd_text = match[1].strip()
    
    # Try to parse date: "31 Dec 2025"
    date_match = re.search(r'(\d{1,2})\s+(\w{3,})\s+(\d{4})', lsd_text)
    if date_match:
        day, month_str, year = date_match.groups()
        month_map = {
            'jan': 1, 'january': 1, 'feb': 2, 'february': 2, 'mar': 3, 'march': 3,
            'apr': 4, 'april': 4, 'may': 5, 'jun': 6, 'june': 6,
            'jul': 7, 'july': 7, 'aug': 8, 'august': 8, 'sep': 9, 'september': 9,
            'oct': 10, 'october': 10, 'nov': 11, 'november': 11, 'dec': 12, 'december': 12
        }
        month = month_map.get(month_str.lower())
        if month:
            return datetime(int(year), month, int(day)), lsd_text
    
    return None, lsd_text


def parse_milestones(text: str) -> List[Dict[str, Any]]:
    """Parse milestones from fee arrangement text"""
    milestones = []
    seen_ordinals = set()  # Track seen ordinals to avoid duplicates
    
    # Split text into lines and process each line
    lines = text.split('\n')
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Pattern: (a) description (50%) - 250,000
        # Use search to find pattern anywhere in line (handles "Original EL: (a) ...")
        match = re.search(
            r'\(([a-z])\)\s*(.+?)(?:\((\d+(?:\.\d+)?)%\))?\s*-\s*([\d,]+)',
            line,
            re.IGNORECASE
        )
        
        if match:
            ordinal, description, percent_str, amount_str = match.groups()
            ordinal_key = f"({ordinal.lower()})"
            
            # Skip if we've seen this ordinal before (avoid duplicates from multi-section ELs)
            if ordinal_key in seen_ordinals:
                continue
            seen_ordinals.add(ordinal_key)
            
            # Clean up description - remove trailing whitespace
            description = description.strip()
            
            # Skip if this looks like a header line
            if len(description) < 3 or description.lower() in ['original el', 'supplemental el']:
                continue
            
            try:
                amount = float(amount_str.replace(',', ''))
                percent = float(percent_str) if percent_str else None
                
                milestones.append({
                    'ordinal': ordinal_key,
                    'title': description[:100],
                    'description': description,
                    'trigger_text': description,
                    'amount_value': amount,
                    'amount_currency': 'USD',
                    'is_percent': percent is not None,
                    'percent_value': percent,
                    'sort_order': ord(ordinal.lower()) - ord('a') + 1
                })
            except ValueError:
                continue
    
    return milestones


def parse_bonuses(text: str) -> List[Dict[str, Any]]:
    """Parse bonus information from Chinese and English text"""
    bonuses = []
    
    # Chinese USD: 不低于X万美元奖金 or X万美元奖金
    match = re.search(r'(?:不低于)?(\d+(?:\.\d+)?)万美元奖金', text)
    if match:
        amount = float(match[1]) * 10000
        bonuses.append({
            'description': match[0],
            'amount_usd': amount,
            'amount_cny': None
        })
    
    # Chinese CNY: 不低于X万人民币奖金
    match = re.search(r'(?:不低于)?(\d+(?:\.\d+)?)万(?:人民币|元)奖金', text)
    if match:
        amount = float(match[1]) * 10000
        bonuses.append({
            'description': match[0],
            'amount_usd': None,
            'amount_cny': amount
        })
    
    # English: bonus of $X,XXX or bonus: $X,XXX
    match = re.search(r'bonus[:\s]+\$\s*([\d,]+)', text, re.IGNORECASE)
    if match:
        amount = float(match[1].replace(',', ''))
        bonuses.append({
            'description': match[0],
            'amount_usd': amount,
            'amount_cny': None
        })
    
    return bonuses


def load_excel_data(file_path: str) -> List[ProjectData]:
    """Load and parse Excel file"""
    print_info(f"Loading Excel file: {file_path}")
    
    if not os.path.exists(file_path):
        print_error(f"Excel file not found: {file_path}")
        sys.exit(1)
    
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
    sheet = wb['Transactions']
    
    projects = []
    
    # Data starts from row 5 (row 4 is headers)
    for row_idx in range(5, sheet.max_row + 1):
        row = sheet[row_idx]
        
        # Skip empty rows
        if not row[2].value:  # Column C - Project Name
            continue
        
        # Get fee arrangement text and strikethrough detection
        # Column I (index 8) contains fee arrangements
        fee_cell = sheet.cell(row=row_idx, column=9)  # Column I
        fee_text, completed = extract_fee_arrangement_text(fee_cell)
        
        # Parse milestones and bonuses
        milestones = parse_milestones(fee_text) if fee_text else []
        bonuses = parse_bonuses(fee_text) if fee_text else []
        
        # Parse LSD
        lsd_date, lsd_raw = parse_lsd(fee_text) if fee_text else (None, None)
        
        project = ProjectData(
            row_num=row_idx,
            project_name=str(row[2].value or "").strip(),  # C
            client_name=str(row[3].value or "").strip(),   # D
            cm_no=str(row[4].value or "").strip(),         # E
            attorney_in_charge=str(row[5].value or "").strip(),  # F
            sca=str(row[6].value or "").strip(),           # G
            fees_usd=parse_money(row[7].value),            # H
            billing_usd=parse_money(row[9].value),         # J (index 9)
            collection_usd=parse_money(row[10].value),     # K
            billing_credit_usd=parse_money(row[11].value), # L
            ubt_usd=parse_money(row[12].value),            # M
            ar_usd=parse_money(row[13].value),             # N
            billing_credit_cny=parse_money(row[15].value), # P
            ubt_cny=parse_money(row[16].value),            # Q
            finance_comment=str(row[17].value or "").strip(),  # R
            remarks=str(row[20].value or "").strip(),      # U
            matter_notes=str(row[21].value or "").strip(), # V
            fee_arrangement_text=fee_text,
            milestones=milestones,
            bonuses=bonuses,
            completed_milestones=completed
        )
        
        # Attach LSD to project for later use
        project.lsd_date = lsd_date
        project.lsd_raw = lsd_raw
        
        projects.append(project)
    
    wb.close()
    print_success(f"Loaded {len(projects)} projects from Excel")
    return projects


class DatabaseUpdater:
    """Handles all database updates"""
    
    def __init__(self, db_url: str):
        self.conn = psycopg2.connect(db_url)
        self.cur = self.conn.cursor(cursor_factory=RealDictCursor)
        self.stats = {
            'financials_updated': 0,
            'milestones_created': 0,
            'milestones_updated': 0,
            'bonuses_updated': 0,
            'comments_added': 0,
            'skipped': 0
        }
    
    def close(self):
        self.cur.close()
        self.conn.close()
    
    def find_cm_by_number(self, cm_no: str) -> Optional[dict]:
        """Find C/M record by number"""
        self.cur.execute(
            "SELECT cm_id, project_id, cm_no FROM billing_project_cm_no WHERE cm_no = %s",
            (cm_no,)
        )
        return self.cur.fetchone()
    
    def find_engagement_by_cm(self, cm_id: int) -> Optional[dict]:
        """Find engagement by C/M ID"""
        self.cur.execute(
            "SELECT engagement_id FROM billing_engagement WHERE cm_id = %s LIMIT 1",
            (cm_id,)
        )
        return self.cur.fetchone()
    
    def update_financials(self, project: ProjectData) -> bool:
        """Update financial data for a project"""
        cm = self.find_cm_by_number(project.cm_no)
        if not cm:
            return False
        
        cm_id = cm['cm_id']
        
        # Build update fields
        updates = []
        params = []
        
        if project.billing_usd > 0:
            updates.append("billing_to_date_usd = %s")
            params.append(project.billing_usd)
        
        if project.collection_usd > 0:
            updates.append("collected_to_date_usd = %s")
            params.append(project.collection_usd)
        
        if project.ubt_usd > 0:
            updates.append("ubt_usd = %s")
            params.append(project.ubt_usd)
        
        if project.billing_credit_usd > 0:
            updates.append("billing_credit_usd = %s")
            params.append(project.billing_credit_usd)
        
        if project.billing_credit_cny > 0:
            updates.append("billing_credit_cny = %s")
            params.append(project.billing_credit_cny)
        
        if project.ubt_cny > 0:
            updates.append("ubt_cny = %s")
            params.append(project.ubt_cny)
        
        if not updates:
            return False
        
        updates.append("financials_updated_at = %s")
        params.append(datetime.utcnow())
        
        params.append(cm_id)
        
        sql = f"UPDATE billing_project_cm_no SET {', '.join(updates)} WHERE cm_id = %s"
        self.cur.execute(sql, params)
        
        if self.cur.rowcount > 0:
            self.stats['financials_updated'] += 1
            return True
        return False
    
    def update_fee_arrangement(self, project: ProjectData) -> Optional[int]:
        """Create or update fee arrangement, return fee_id"""
        cm = self.find_cm_by_number(project.cm_no)
        if not cm:
            return None
        
        engagement = self.find_engagement_by_cm(cm['cm_id'])
        if not engagement:
            return None
        
        engagement_id = engagement['engagement_id']
        
        # Check if fee arrangement exists
        self.cur.execute(
            "SELECT fee_id FROM billing_fee_arrangement WHERE engagement_id = %s",
            (engagement_id,)
        )
        existing = self.cur.fetchone()
        
        if existing:
            fee_id = existing['fee_id']
            # Update existing
            self.cur.execute("""
                UPDATE billing_fee_arrangement 
                SET raw_text = %s, lsd_date = %s, lsd_raw = %s, updated_at = NOW()
                WHERE fee_id = %s
            """, (project.fee_arrangement_text, project.lsd_date, project.lsd_raw, fee_id))
        else:
            # Create new
            self.cur.execute("""
                INSERT INTO billing_fee_arrangement (engagement_id, raw_text, lsd_date, lsd_raw, parsed_at)
                VALUES (%s, %s, %s, %s, NOW())
                RETURNING fee_id
            """, (engagement_id, project.fee_arrangement_text, project.lsd_date, project.lsd_raw))
            fee_id = self.cur.fetchone()['fee_id']
        
        return fee_id
    
    def update_milestones(self, fee_id: int, project: ProjectData) -> int:
        """Update milestones for a fee arrangement"""
        if not project.milestones:
            return 0
        
        count = 0
        for milestone in project.milestones:
            # Check if milestone exists
            self.cur.execute(
                "SELECT milestone_id, completed FROM billing_milestone WHERE fee_id = %s AND ordinal = %s",
                (fee_id, milestone['ordinal'])
            )
            existing = self.cur.fetchone()
            
            if existing:
                milestone_id = existing['milestone_id']
                was_completed = existing['completed']
                
                # Update existing
                self.cur.execute("""
                    UPDATE billing_milestone 
                    SET title = %s, description = %s, amount_value = %s, 
                        is_percent = %s, percent_value = %s, sort_order = %s,
                        updated_at = NOW()
                    WHERE milestone_id = %s
                """, (
                    milestone['title'], milestone['description'], milestone['amount_value'],
                    milestone['is_percent'], milestone['percent_value'], milestone['sort_order'],
                    milestone_id
                ))
                
                # Check if this milestone should be marked completed (from strikethrough)
                if milestone['ordinal'] in project.completed_milestones and not was_completed:
                    self.cur.execute("""
                        UPDATE billing_milestone 
                        SET completed = true, completion_source = 'excel_strikethrough', 
                            completion_date = CURRENT_DATE, updated_at = NOW()
                        WHERE milestone_id = %s
                    """, (milestone_id,))
                    self.stats['milestones_updated'] += 1
                
            else:
                # Create new
                self.cur.execute("""
                    INSERT INTO billing_milestone 
                    (fee_id, ordinal, title, description, trigger_type, trigger_text,
                     amount_value, amount_currency, is_percent, percent_value, sort_order,
                     completed, raw_fragment)
                    VALUES (%s, %s, %s, %s, 'date_based', %s, %s, 'USD', %s, %s, %s, %s, %s)
                    RETURNING milestone_id
                """, (
                    fee_id, milestone['ordinal'], milestone['title'], milestone['description'],
                    milestone['trigger_text'], milestone['amount_value'],
                    milestone['is_percent'], milestone['percent_value'], milestone['sort_order'],
                    milestone['ordinal'] in project.completed_milestones,
                    milestone['description']
                ))
                milestone_id = self.cur.fetchone()['milestone_id']
                
                if milestone['ordinal'] in project.completed_milestones:
                    self.cur.execute("""
                        UPDATE billing_milestone 
                        SET completion_source = 'excel_strikethrough', completion_date = CURRENT_DATE
                        WHERE milestone_id = %s
                    """, (milestone_id,))
                
                self.stats['milestones_created'] += 1
            
            count += 1
        
        return count
    
    def update_bonus(self, fee_id: int, project: ProjectData) -> bool:
        """Update bonus information"""
        if not project.bonuses:
            return False
        
        bonus = project.bonuses[0]  # Take first bonus
        
        self.cur.execute("""
            UPDATE billing_fee_arrangement 
            SET bonus_description = %s, bonus_amount_usd = %s, bonus_amount_cny = %s, updated_at = NOW()
            WHERE fee_id = %s
        """, (bonus['description'], bonus['amount_usd'], bonus['amount_cny'], fee_id))
        
        if self.cur.rowcount > 0:
            self.stats['bonuses_updated'] += 1
            return True
        return False
    
    def add_finance_comment(self, project: ProjectData) -> bool:
        """Add finance comment if present"""
        if not project.finance_comment:
            return False
        
        cm = self.find_cm_by_number(project.cm_no)
        if not cm:
            return False
        
        engagement = self.find_engagement_by_cm(cm['cm_id'])
        if not engagement:
            return False
        
        engagement_id = engagement['engagement_id']
        fingerprint = hashlib.md5(project.finance_comment.encode()).hexdigest()
        
        # Check if comment already exists
        self.cur.execute(
            "SELECT 1 FROM billing_finance_comment WHERE engagement_id = %s AND fingerprint_hash = %s",
            (engagement_id, fingerprint)
        )
        if self.cur.fetchone():
            return False
        
        self.cur.execute("""
            INSERT INTO billing_finance_comment (engagement_id, comment_raw, fingerprint_hash, parsed_at)
            VALUES (%s, %s, %s, NOW())
        """, (engagement_id, project.finance_comment, fingerprint))
        
        self.stats['comments_added'] += 1
        return True
    
    def commit(self):
        self.conn.commit()


def main():
    print_header("BILLING DATABASE UPDATE")
    
    # Check Excel file
    if not os.path.exists(EXCEL_FILE):
        print_error(f"Excel file not found: {EXCEL_FILE}")
        print_info("Set EXCEL_FILE environment variable or update DEFAULT_EXCEL_FILE in script")
        sys.exit(1)
    
    print_info(f"Excel file: {EXCEL_FILE}")
    print_info(f"Database: {DATABASE_URL.split('@')[1] if '@' in DATABASE_URL else 'configured'}")
    
    # Load Excel data
    projects = load_excel_data(EXCEL_FILE)
    
    if not projects:
        print_error("No projects found in Excel file")
        sys.exit(1)
    
    # Connect to database
    print_info("Connecting to database...")
    try:
        updater = DatabaseUpdater(DATABASE_URL)
    except Exception as e:
        print_error(f"Failed to connect to database: {e}")
        sys.exit(1)
    
    try:
        print_header("UPDATING DATABASE")
        
        for i, project in enumerate(projects, 1):
            print(f"\nProcessing {i}/{len(projects)}: {project.project_name[:50]}", end=" ")
            
            # 1. Update financials
            if updater.update_financials(project):
                print("[F]", end="")  # Financials updated
            
            # 2. Update fee arrangement and milestones
            if project.fee_arrangement_text:
                fee_id = updater.update_fee_arrangement(project)
                if fee_id:
                    count = updater.update_milestones(fee_id, project)
                    if count > 0:
                        print(f"[M{count}]", end="")  # Milestones
                    
                    if updater.update_bonus(fee_id, project):
                        print("[B]", end="")  # Bonus
            
            # 3. Add finance comment
            if updater.add_finance_comment(project):
                print("[C]", end="")  # Comment
            
            # Show progress every 10 projects
            if i % 10 == 0:
                print(f" ({i}/{len(projects)})")
        
        # Commit all changes
        updater.commit()
        
        # Print summary
        print_header("UPDATE SUMMARY")
        print_success(f"Financial records updated: {updater.stats['financials_updated']}")
        print_success(f"Milestones created: {updater.stats['milestones_created']}")
        print_success(f"Milestones marked completed: {updater.stats['milestones_updated']}")
        print_success(f"Bonuses updated: {updater.stats['bonuses_updated']}")
        print_success(f"Finance comments added: {updater.stats['comments_added']}")
        
        print("\n" + "="*80)
        print("✅ Billing database update completed successfully!")
        print("="*80)
        
    except Exception as e:
        print_error(f"Error during update: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        updater.close()


if __name__ == "__main__":
    main()
