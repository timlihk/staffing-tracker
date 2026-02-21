#!/usr/bin/env python3
"""
Simple synchronous billing updater - shows progress for each project
"""

import os
import sys
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime

import openpyxl
import psycopg2

DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://postgres:qtSnaaSaelqHVydazTmViwejXbkkZxVY@crossover.proxy.rlwy.net:15782/railway",
)

EXCEL_FILE = "/Users/timli/Library/CloudStorage/OneDrive-Personal/Coding/staffing-tracker/Billing/HKCM Project List (2026.02.12).xlsx"
NS = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}


def parse_money(value):
    if value is None:
        return 0.0
    s = str(value).strip().replace(",", "").replace("$", "").replace("，", "")
    if s in ("", "-", "—", "–"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_milestones(text, struck_ordinals):
    """Simple milestone parser"""
    if not text:
        return []
    
    milestones = []
    seen = set()
    
    for match in re.finditer(r'\(([a-z])\)', text, re.I):
        ordinal = f"({match.group(1).lower()})"
        if ordinal in seen:
            continue
        seen.add(ordinal)
        
        start = match.end()
        # Find next ordinal or end
        next_match = None
        for m2 in re.finditer(r'\(([a-z])\)', text[start:], re.I):
            next_match = m2
            break
        
        if next_match:
            segment = text[start:start + next_match.start()].strip()
        else:
            segment = text[start:].strip()
        
        # Clean segment
        segment = re.sub(r'\n.*$', '', segment, flags=re.DOTALL)
        
        # Extract amount
        amount = None
        amt_match = re.search(r'[-–—]?\s*([\d,]+(?:\.\d+)?)', segment.replace(',', ''))
        if amt_match:
            try:
                val = float(amt_match.group(1))
                if val > 1000:
                    amount = val
            except:
                pass
        
        completed = ordinal in struck_ordinals or "*" in struck_ordinals
        
        milestones.append({
            'ordinal': ordinal,
            'description': segment[:200],
            'amount': amount,
            'currency': 'USD',
            'completed': completed
        })
    
    return milestones


def main():
    print("="*80)
    print("KIMI BILLING DATABASE UPDATER - SIMPLE")
    print("="*80)
    
    # Load Excel
    print("\n1. Loading Excel...")
    wb_data = openpyxl.load_workbook(EXCEL_FILE, data_only=True, read_only=True)
    sheet_data = wb_data['Transactions']
    
    wb_format = openpyxl.load_workbook(EXCEL_FILE, data_only=False, read_only=True)
    sheet_format = wb_format['Transactions']
    print(f"   Sheet has {sheet_data.max_row} rows")
    
    # Load strikethrough data from XML
    print("\n2. Loading strikethrough data...")
    struck_data = {}
    with zipfile.ZipFile(EXCEL_FILE, 'r') as zf:
        ss_xml = zf.read('xl/sharedStrings.xml')
        root = ET.fromstring(ss_xml)
        
        for idx, si in enumerate(root.findall('.//main:si', NS)):
            for r in si.findall('.//main:r', NS):
                rpr = r.find('.//main:rPr', NS)
                if rpr is not None and rpr.find('.//main:strike', NS) is not None:
                    t = r.find('.//main:t', NS)
                    if t is not None and t.text:
                        text = t.text
                        for m in re.finditer(r'\(([a-z])\)', text, re.I):
                            if idx not in struck_data:
                                struck_data[idx] = set()
                            struck_data[idx].add(f"({m[1].lower()})")
    
    print(f"   Found {len(struck_data)} cells with strikethrough")
    
    # Connect to database
    print("\n3. Connecting to database...")
    conn = psycopg2.connect(DATABASE_URL)
    conn.autocommit = False
    cur = conn.cursor()
    print("   Connected!")
    
    # Load CM cache
    print("\n4. Loading CM cache...")
    cur.execute("SELECT cm_no, project_id FROM billing_project_cm_no")
    cm_cache = {row[0]: row[1] for row in cur.fetchall()}
    print(f"   Cached {len(cm_cache)} CM numbers")
    
    # Process rows
    print("\n5. Processing projects...")
    print("-"*80)
    
    stats = {'matched': 0, 'not_found': 0, 'milestones': 0}
    last_cm_no = ""
    batch_count = 0
    
    for row_idx in range(5, sheet_data.max_row + 1):
        row = [cell.value for cell in sheet_data[row_idx]]
        
        if len(row) < 9:
            continue
        
        project_name = str(row[1]) if row[1] else ""
        client_name = str(row[2]) if row[2] else ""
        cm_no = str(row[3]) if row[3] else ""
        attorney = str(row[4]) if row[4] else ""
        sca = str(row[5]) if row[5] else ""
        
        # Handle empty CM
        if not cm_no or cm_no.strip().lower() in ("", "none"):
            if last_cm_no:
                cm_no = last_cm_no
            else:
                continue
        else:
            last_cm_no = cm_no
        
        fee_text = str(row[8]) if row[8] else ""
        if not fee_text or len(fee_text) < 10:
            continue
        
        # Get strikethrough
        cell_strike = False
        try:
            cell = sheet_format.cell(row_idx, 9)
            if cell.font:
                cell_strike = cell.font.strike or False
        except:
            pass
        
        struck_ordinals = set()
        if cell_strike:
            struck_ordinals.add("*")
        
        # Parse milestones
        milestones = parse_milestones(fee_text, struck_ordinals)
        
        # Find project
        project_id = cm_cache.get(cm_no)
        if not project_id:
            # Try partial match
            for cached_cm, pid in cm_cache.items():
                if cm_no in cached_cm or cached_cm in cm_no:
                    project_id = pid
                    break
        
        if not project_id:
            stats['not_found'] += 1
            if stats['not_found'] <= 3:
                print(f"   Row {row_idx}: NOT FOUND - CM={cm_no[:30]}")
            continue
        
        stats['matched'] += 1
        stats['milestones'] += len(milestones)
        
        # Show progress every 10 projects
        if stats['matched'] % 10 == 0:
            print(f"   Progress: {stats['matched']} matched, {stats['not_found']} not found, {stats['milestones']} milestones")
        
        # Update project
        cur.execute("""
            UPDATE billing_project SET
                client_name = %s,
                attorney_in_charge = %s,
                sca = %s,
                updated_at = NOW()
            WHERE project_id = %s
        """, (client_name, attorney, sca, project_id))
        
        # Get or create CM
        cur.execute("SELECT cm_id FROM billing_project_cm_no WHERE cm_no = %s", (cm_no,))
        result = cur.fetchone()
        
        if result:
            cm_id = result[0]
        else:
            cur.execute("""
                INSERT INTO billing_project_cm_no (project_id, cm_no, is_primary, status)
                VALUES (%s, %s, FALSE, 'active')
                RETURNING cm_id
            """, (project_id, cm_no))
            cm_id = cur.fetchone()[0]
            cm_cache[cm_no] = project_id
        
        # Get or create engagement
        cur.execute("SELECT engagement_id FROM billing_engagement WHERE cm_id = %s", (cm_id,))
        result = cur.fetchone()
        
        if result:
            engagement_id = result[0]
        else:
            cur.execute("""
                INSERT INTO billing_engagement (project_id, cm_id, engagement_code, engagement_title, updated_at)
                VALUES (%s, %s, 'original', %s, NOW())
                RETURNING engagement_id
            """, (project_id, cm_id, project_name[:100]))
            engagement_id = cur.fetchone()[0]
        
        # Get or create fee arrangement
        cur.execute("SELECT fee_id FROM billing_fee_arrangement WHERE engagement_id = %s", (engagement_id,))
        result = cur.fetchone()
        
        if result:
            fee_id = result[0]
            cur.execute("""
                UPDATE billing_fee_arrangement 
                SET raw_text = %s, parser_version = 'kimi-v1', parsed_at = NOW(), updated_at = NOW()
                WHERE fee_id = %s
            """, (fee_text, fee_id))
        else:
            cur.execute("""
                INSERT INTO billing_fee_arrangement (engagement_id, raw_text, parser_version, parsed_at, created_at, updated_at)
                VALUES (%s, %s, 'kimi-v1', NOW(), NOW(), NOW())
                RETURNING fee_id
            """, (engagement_id, fee_text))
            fee_id = cur.fetchone()[0]
        
        # Upsert milestones
        for i, m in enumerate(milestones):
            cur.execute("""
                SELECT milestone_id FROM billing_milestone
                WHERE engagement_id = %s AND ordinal = %s
            """, (engagement_id, m['ordinal']))
            result = cur.fetchone()
            
            if result:
                cur.execute("""
                    UPDATE billing_milestone SET
                        description = %s,
                        amount_value = %s,
                        amount_currency = %s,
                        completed = %s,
                        completion_source = CASE WHEN %s THEN 'excel_strikethrough' ELSE completion_source END,
                        updated_at = NOW(),
                        sort_order = %s,
                        raw_fragment = %s
                    WHERE milestone_id = %s
                """, (
                    m['description'], m['amount'], m['currency'], m['completed'],
                    m['completed'], i, m['description'][:100], result[0]
                ))
            else:
                cur.execute("""
                    INSERT INTO billing_milestone (
                        engagement_id, fee_id, ordinal, description,
                        amount_value, amount_currency, completed,
                        completion_source, created_at, updated_at, sort_order, raw_fragment
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW(), %s, %s)
                """, (
                    engagement_id, fee_id, m['ordinal'], m['description'],
                    m['amount'], m['currency'], m['completed'],
                    'excel_strikethrough' if m['completed'] else None,
                    i, m['description'][:100]
                ))
        
        # Commit every 5 projects
        batch_count += 1
        if batch_count >= 5:
            conn.commit()
            batch_count = 0
    
    # Final commit
    conn.commit()
    
    print("-"*80)
    print("\n✅ COMPLETE!")
    print(f"   Projects matched: {stats['matched']}")
    print(f"   Projects not found: {stats['not_found']}")
    print(f"   Total milestones: {stats['milestones']}")
    
    cur.close()
    conn.close()
    wb_data.close()
    wb_format.close()


if __name__ == "__main__":
    main()
