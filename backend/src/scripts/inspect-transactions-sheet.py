#!/usr/bin/env python3
"""
Inspect the Transactions worksheet structure
"""

import openpyxl

EXCEL_FILE = '/home/timlihk/staffing-tracker/billing-matter/HKCM Project List(81764217.1)_6Oct25.xlsx'

print("Loading Excel file (fast mode)...")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True, read_only=True)
sheet = wb['Transactions']

print(f"Transactions worksheet: {sheet.max_row} rows, {sheet.max_column} columns\n")

# Print first 10 rows
print("First 10 rows:")
for row_idx in range(1, min(11, sheet.max_row + 1)):
    print(f"\n  Row {row_idx}:")
    row_data = []
    for col_idx, cell in enumerate(sheet[row_idx], start=1):
        if cell.value is not None and str(cell.value).strip():
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            value = str(cell.value).strip()[:50]
            row_data.append(f"{col_letter}:{value}")

    if row_data:
        print(f"    {' | '.join(row_data)}")
    else:
        print("    (empty)")

wb.close()
