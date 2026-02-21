#!/usr/bin/env python3
"""
AI-Powered Milestone Parser

Uses LLM to intelligently parse milestone text from Excel fee arrangements,
handling various formats including:
- (a), (b), (c) or (1), (2), (3) ordinals
- Mixed Chinese/English text
- Various separators and formatting
- Strikethrough detection for completion status
- Dollar amount extraction per milestone

Usage:
    export OPENAI_API_KEY="your-key"
    export DATABASE_URL="your-db-url"
    python3 src/scripts/ai-parse-milestones.py [--dry-run] [--project PROJECT_NAME]
"""

import os
import sys
import json
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
from typing import Dict, List, Optional, Set, Tuple, Any
from dataclasses import dataclass, asdict
from textwrap import dedent

import openpyxl
import psycopg2
from psycopg2.extras import RealDictCursor

# Try to import OpenAI
try:
    from openai import OpenAI
    HAS_OPENAI = True
except ImportError:
    HAS_OPENAI = False
    print("Warning: openai package not installed. Install with: pip install openai")

# Configuration
DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://postgres:qtSnaaSaelqHVydazTmViwejXbkkZxVY@crossover.proxy.rlwy.net:15782/railway",
)

DEFAULT_EXCEL_FILE = "/Users/timli/Library/CloudStorage/OneDrive-Personal/Coding/staffing-tracker/Billing/HKCM Project List (2026.02.12).xlsx"
EXCEL_FILE = os.environ.get("EXCEL_FILE", DEFAULT_EXCEL_FILE)

OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")

# XML namespaces for Excel
NS = {
    'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
}


@dataclass
class Milestone:
    """Represents a parsed milestone"""
    ordinal: str  # e.g., "(a)", "(1)"
    title: str
    description: str
    amount_usd: Optional[float]
    amount_cny: Optional[float]
    percentage: Optional[float]
    is_conditional: bool  # True if no fixed amount (calculated/adjusted)
    completed: bool
    raw_text: str  # Original text for this milestone
    
    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


@dataclass
class FeeArrangement:
    """Represents a parsed fee arrangement"""
    project_name: str
    cm_no: str
    row_num: int
    lsd_date: Optional[str]
    lsd_raw: Optional[str]
    milestones: List[Milestone]
    bonus_usd: Optional[float]
    bonus_cny: Optional[float]
    bonus_description: Optional[str]
    raw_text: str
    strikethrough_ordinals: Set[str]


class StrikethroughExtractor:
    """Extract strikethrough formatting from Excel XML"""
    
    @staticmethod
    def extract(excel_path: str, sheet_name: str = 'Transactions') -> Dict[int, Set[str]]:
        """Extract strikethrough ordinals per row"""
        strikethrough_data = {}
        
        try:
            with zipfile.ZipFile(excel_path, 'r') as zf:
                # Read shared strings
                shared_strings = []
                if 'xl/sharedStrings.xml' in zf.namelist():
                    ss_xml = zf.read('xl/sharedStrings.xml')
                    root = ET.fromstring(ss_xml)
                    
                    for si in root.findall('.//main:si', NS):
                        text_parts = []
                        has_strike = False
                        
                        for r in si.findall('.//main:r', NS):
                            t = r.find('.//main:t', NS)
                            text = t.text if t is not None else ""
                            
                            rpr = r.find('.//main:rPr', NS)
                            is_struck = False
                            if rpr is not None:
                                strike = rpr.find('.//main:strike', NS)
                                if strike is not None:
                                    val = strike.get('val', 'true')
                                    if val in ('true', '1', 'single'):
                                        is_struck = True
                                        has_strike = True
                            
                            text_parts.append({'text': text, 'struck': is_struck})
                        
                        if not text_parts:
                            t = si.find('.//main:t', NS)
                            if t is not None:
                                text_parts.append({'text': t.text, 'struck': False})
                        
                        shared_strings.append({
                            'parts': text_parts,
                            'has_strike': has_strike
                        })
                
                # Read worksheet
                ws_xml = zf.read('xl/worksheets/sheet1.xml')
                root = ET.fromstring(ws_xml)
                
                for row in root.findall('.//main:row', NS):
                    row_num = int(row.get('r', 0))
                    
                    for cell in row.findall('.//main:c', NS):
                        cell_ref = cell.get('r', '')
                        cell_type = cell.get('t', '')
                        
                        if not cell_ref.startswith('I'):
                            continue
                        
                        if cell_type == 's':
                            v = cell.find('.//main:v', NS)
                            if v is not None:
                                try:
                                    ss_idx = int(v.text)
                                    if ss_idx < len(shared_strings):
                                        ss_item = shared_strings[ss_idx]
                                        if ss_item['has_strike']:
                                            struck_ordinals = set()
                                            for part in ss_item['parts']:
                                                if part['struck']:
                                                    for match in re.finditer(r'\(([a-z0-9])\)', part['text'], re.IGNORECASE):
                                                        struck_ordinals.add(f"({match[1].lower()})")
                                            
                                            if struck_ordinals:
                                                strikethrough_data[row_num] = struck_ordinals
                                except (ValueError, IndexError):
                                    pass
        except Exception as e:
            print(f"Warning: Could not extract strikethrough: {e}")
        
        return strikethrough_data


class AIMilestoneParser:
    """Uses LLM to parse milestone text"""
    
    def __init__(self):
        self.client = None
        if HAS_OPENAI and OPENAI_API_KEY:
            self.client = OpenAI(api_key=OPENAI_API_KEY)
    
    def parse_milestones(self, text: str, strikethrough_ordinals: Set[str]) -> List[Milestone]:
        """Parse milestone text using AI"""
        if not self.client:
            print("Error: OpenAI client not available")
            return []
        
        system_prompt = dedent("""
            You are a legal billing expert. Parse the following fee arrangement text into structured milestones.
            
            Rules:
            1. Identify each milestone by its ordinal: (a), (b), (c), (d), (e), (f), etc. OR (1), (2), (3), etc.
            2. Extract the description for each milestone
            3. Extract the dollar amount (USD) if present. If no amount but has percentage, note it.
            4. If the milestone mentions calculated/adjusted/diff amounts, mark as conditional (no fixed amount)
            5. The text may be in Chinese, English, or mixed
            6. Look for LSD (Long Stop Date) - usually in format "(LSD: DD MMM YYYY)"
            7. Look for bonus amounts mentioned
            8. Return ONLY valid JSON
            
            Output format:
            {
                "lsd_date": "YYYY-MM-DD or null",
                "lsd_raw": "original LSD text or null",
                "milestones": [
                    {
                        "ordinal": "(a)",
                        "title": "short description",
                        "description": "full description",
                        "amount_usd": 95000 or null,
                        "amount_cny": null,
                        "percentage": 50 or null,
                        "is_conditional": false,
                        "raw_text": "original text for this milestone"
                    }
                ],
                "bonus_usd": null or number,
                "bonus_cny": null or number,
                "bonus_description": null or string
            }
        """)
        
        user_prompt = f"Parse this fee arrangement text:\n\n{text}\n\nStrikethrough ordinals (completed): {list(strikethrough_ordinals)}"
        
        try:
            response = self.client.chat.completions.create(
                model="gpt-4o-mini",  # Use mini for cost-effectiveness
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt}
                ],
                temperature=0.1,  # Low temperature for consistent output
                response_format={"type": "json_object"}
            )
            
            result = json.loads(response.choices[0].message.content)
            
            # Convert to Milestone objects
            milestones = []
            for m in result.get('milestones', []):
                ordinal = m.get('ordinal', '')
                # Mark as completed if in strikethrough list
                completed = ordinal in strikethrough_ordinals
                
                milestones.append(Milestone(
                    ordinal=ordinal,
                    title=m.get('title', '')[:100],
                    description=m.get('description', ''),
                    amount_usd=m.get('amount_usd'),
                    amount_cny=m.get('amount_cny'),
                    percentage=m.get('percentage'),
                    is_conditional=m.get('is_conditional', False),
                    completed=completed,
                    raw_text=m.get('raw_text', '')
                ))
            
            return milestones, result
            
        except Exception as e:
            print(f"AI parsing error: {e}")
            return [], {}


class DatabaseUpdater:
    """Updates database with parsed milestones"""
    
    def __init__(self, db_url: str):
        self.conn = psycopg2.connect(db_url)
        self.cur = self.conn.cursor(cursor_factory=RealDictCursor)
        self.stats = {
            'milestones_created': 0,
            'milestones_updated': 0,
            'fee_arrangements_updated': 0,
            'errors': 0
        }
    
    def close(self):
        self.cur.close()
        self.conn.close()
    
    def find_cm_by_number(self, cm_no: str) -> Optional[dict]:
        self.cur.execute(
            "SELECT cm_id, project_id FROM billing_project_cm_no WHERE cm_no = %s",
            (cm_no,)
        )
        return self.cur.fetchone()
    
    def update_fee_arrangement(self, fee: FeeArrangement) -> Optional[int]:
        """Create/update fee arrangement and milestones"""
        try:
            cm = self.find_cm_by_number(fee.cm_no)
            if not cm:
                print(f"  ⚠️ C/M not found: {fee.cm_no}")
                return None
            
            cm_id = cm['cm_id']
            
            # Find engagement
            self.cur.execute(
                "SELECT engagement_id FROM billing_engagement WHERE cm_id = %s",
                (cm_id,)
            )
            engagement = self.cur.fetchone()
            if not engagement:
                print(f"  ⚠️ Engagement not found for C/M {fee.cm_no}")
                return None
            
            engagement_id = engagement['engagement_id']
            
            # Check if fee arrangement exists
            self.cur.execute(
                "SELECT fee_id FROM billing_fee_arrangement WHERE engagement_id = %s",
                (engagement_id,)
            )
            existing = self.cur.fetchone()
            
            if existing:
                fee_id = existing['fee_id']
                # Update
                self.cur.execute("""
                    UPDATE billing_fee_arrangement 
                    SET raw_text = %s, lsd_date = %s, lsd_raw = %s,
                        bonus_description = %s, bonus_amount_usd = %s, bonus_amount_cny = %s,
                        updated_at = NOW()
                    WHERE fee_id = %s
                """, (
                    fee.raw_text, fee.lsd_date, fee.lsd_raw,
                    fee.bonus_description, fee.bonus_usd, fee.bonus_cny,
                    fee_id
                ))
            else:
                # Create new
                self.cur.execute("""
                    INSERT INTO billing_fee_arrangement 
                    (engagement_id, raw_text, lsd_date, lsd_raw, 
                     bonus_description, bonus_amount_usd, bonus_amount_cny, parsed_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
                    RETURNING fee_id
                """, (
                    engagement_id, fee.raw_text, fee.lsd_date, fee.lsd_raw,
                    fee.bonus_description, fee.bonus_usd, fee.bonus_cny
                ))
                fee_id = self.cur.fetchone()['fee_id']
            
            # Process milestones
            for milestone in fee.milestones:
                self._upsert_milestone(fee_id, milestone)
            
            self.conn.commit()
            return fee_id
            
        except Exception as e:
            self.conn.rollback()
            print(f"  ❌ Error: {e}")
            self.stats['errors'] += 1
            return None
    
    def _upsert_milestone(self, fee_id: int, milestone: Milestone):
        """Create or update a milestone"""
        try:
            # Check if exists
            self.cur.execute("""
                SELECT milestone_id, completed FROM billing_milestone 
                WHERE fee_id = %s AND ordinal = %s
            """, (fee_id, milestone.ordinal))
            
            existing = self.cur.fetchone()
            
            if existing:
                milestone_id = existing['milestone_id']
                was_completed = existing['completed']
                
                # Update
                self.cur.execute("""
                    UPDATE billing_milestone 
                    SET title = %s, description = %s, amount_value = %s,
                        amount_currency = %s, is_percent = %s, percent_value = %s,
                        updated_at = NOW()
                    WHERE milestone_id = %s
                """, (
                    milestone.title, milestone.description, milestone.amount_usd,
                    'USD' if milestone.amount_usd else None,
                    milestone.percentage is not None, milestone.percentage,
                    milestone_id
                ))
                
                # Update completion if changed
                if milestone.completed and not was_completed:
                    self.cur.execute("""
                        UPDATE billing_milestone 
                        SET completed = true, completion_source = 'excel_strikethrough',
                            completion_date = CURRENT_DATE, updated_at = NOW()
                        WHERE milestone_id = %s
                    """, (milestone_id,))
                    self.stats['milestones_updated'] += 1
            else:
                # Create new
                self.cur.execute("""
                    INSERT INTO billing_milestone 
                    (fee_id, ordinal, title, description, trigger_type, trigger_text,
                     amount_value, amount_currency, is_percent, percent_value,
                     completed, completion_source, completion_date, raw_fragment, sort_order)
                    VALUES (%s, %s, %s, %s, 'date_based', %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING milestone_id
                """, (
                    fee_id, milestone.ordinal, milestone.title, milestone.description,
                    milestone.description, milestone.amount_usd,
                    'USD' if milestone.amount_usd else None,
                    milestone.percentage is not None, milestone.percentage,
                    milestone.completed,
                    'excel_strikethrough' if milestone.completed else None,
                    datetime.now().date() if milestone.completed else None,
                    milestone.raw_text,
                    ord(milestone.ordinal[1].lower()) - ord('a') + 1 if milestone.ordinal[1].isalpha() else int(milestone.ordinal[1])
                ))
                self.stats['milestones_created'] += 1
                
        except Exception as e:
            print(f"    Error upserting milestone {milestone.ordinal}: {e}")
            raise


def main():
    import argparse
    parser = argparse.ArgumentParser(description='AI-powered milestone parser')
    parser.add_argument('--dry-run', action='store_true', help='Parse only, do not update DB')
    parser.add_argument('--project', type=str, help='Process specific project only')
    parser.add_argument('--limit', type=int, help='Limit number of projects to process')
    args = parser.parse_args()
    
    print("=" * 80)
    print("AI-POWERED MILESTONE PARSER")
    print("=" * 80)
    
    if not HAS_OPENAI:
        print("\n❌ Error: OpenAI package not installed")
        print("   Install with: pip install openai")
        sys.exit(1)
    
    if not OPENAI_API_KEY:
        print("\n❌ Error: OPENAI_API_KEY not set")
        print("   Set with: export OPENAI_API_KEY='your-key'")
        sys.exit(1)
    
    # Extract strikethrough data
    print("\nExtracting strikethrough formatting...")
    strikethrough_data = StrikethroughExtractor.extract(EXCEL_FILE)
    print(f"Found {len(strikethrough_data)} rows with strikethrough")
    
    # Load Excel
    print("\nLoading Excel...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    sheet = wb['Transactions']
    
    # Initialize AI parser
    ai_parser = AIMilestoneParser()
    
    # Process rows
    fees_to_process = []
    
    for row_idx in range(5, sheet.max_row + 1):
        project_name = sheet.cell(row_idx, 3).value
        if not project_name:
            continue
        
        # Filter by project name if specified
        if args.project and args.project.lower() not in str(project_name).lower():
            continue
        
        cm_no = str(sheet.cell(row_idx, 5).value or "").strip()
        fee_text = str(sheet.cell(row_idx, 9).value or "")
        
        if not fee_text or len(fee_text) < 10:
            continue
        
        strikethrough_ordinals = strikethrough_data.get(row_idx, set())
        
        fees_to.append({
            'row': row_idx,
            'project': project_name,
            'cm_no': cm_no,
            'text': fee_text,
            'completed': strikethrough_ordinals
        })
    
    wb.close()
    
    if args.limit:
        fees_to_process = fees_to_process[:args.limit]
    
    print(f"\nProcessing {len(fees_to_process)} fee arrangements...")
    
    # Parse with AI
    parsed_fees = []
    
    for i, item in enumerate(fees_to_process, 1):
        print(f"\n[{i}/{len(fees_to_process)}] {item['project']} (Row {item['row']})")
        print(f"   C/M: {item['cm_no']}")
        print(f"   Strikethrough ordinals: {item['completed']}")
        
        milestones, result = ai_parser.parse_milestones(item['text'], item['completed'])
        
        if milestones:
            print(f"   ✓ Parsed {len(milestones)} milestones:")
            for m in milestones:
                status = "✅" if m.completed else "⏳"
                amt = f"${m.amount_usd:,.0f}" if m.amount_usd else f"{m.percentage}%" if m.percentage else "conditional"
                print(f"      {status} {m.ordinal} {m.title[:50]}... ({amt})")
            
            fee = FeeArrangement(
                project_name=item['project'],
                cm_no=item['cm_no'],
                row_num=item['row'],
                lsd_date=result.get('lsd_date'),
                lsd_raw=result.get('lsd_raw'),
                milestones=milestones,
                bonus_usd=result.get('bonus_usd'),
                bonus_cny=result.get('bonus_cny'),
                bonus_description=result.get('bonus_description'),
                raw_text=item['text'],
                strikethrough_ordinals=item['completed']
            )
            parsed_fees.append(fee)
        else:
            print(f"   ⚠️ No milestones parsed")
    
    # Update database
    if args.dry_run:
        print("\n" + "=" * 80)
        print("DRY RUN - No database updates")
        print("=" * 80)
    else:
        print("\n" + "=" * 80)
        print("UPDATING DATABASE")
        print("=" * 80)
        
        updater = DatabaseUpdater(DATABASE_URL)
        
        try:
            for fee in parsed_fees:
                print(f"\n{fee.project_name}:")
                updater.update_fee_arrangement(fee)
            
            print("\n" + "=" * 80)
            print("SUMMARY")
            print("=" * 80)
            print(f"Milestones created: {updater.stats['milestones_created']}")
            print(f"Milestones updated: {updater.stats['milestones_updated']}")
            print(f"Errors: {updater.stats['errors']}")
            
        finally:
            updater.close()
    
    print("\n✅ Done!")


if __name__ == "__main__":
    main()
