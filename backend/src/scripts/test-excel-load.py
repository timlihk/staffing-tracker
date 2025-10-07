#!/usr/bin/env python3
"""
Test script to check if Excel file can be loaded
"""

import openpyxl
import time

EXCEL_FILE = '/home/timlihk/staffing-tracker/billing-matter/HKCM Project List(81764217.1)_6Oct25.xlsx'

print("Testing Excel file loading...")
print(f"File: {EXCEL_FILE}")

# Test 1: Load with minimal options
print("\nTest 1: Loading with data_only=True (fastest, no formatting)")
start = time.time()
try:
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True, read_only=True)
    print(f"✅ Loaded in {time.time() - start:.2f}s")
    print(f"   Worksheets: {wb.sheetnames}")
    wb.close()
except Exception as e:
    print(f"❌ Failed: {e}")

# Test 2: Load with formatting
print("\nTest 2: Loading with data_only=False (slow, includes formatting)")
start = time.time()
try:
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False, keep_vba=False)
    print(f"✅ Loaded in {time.time() - start:.2f}s")
    print(f"   Worksheets: {wb.sheetnames}")

    # Check first sheet
    sheet = wb.active
    print(f"   Active sheet: {sheet.title}")
    print(f"   Max row: {sheet.max_row}, Max col: {sheet.max_column}")

    wb.close()
except Exception as e:
    print(f"❌ Failed: {e}")

print("\nDone!")
